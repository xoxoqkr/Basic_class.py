# -*- coding: utf-8 -*-
import math
import numpy as np
from openpyxl import Workbook
import random
import csv

data = [[0,25,25,0,0]]
customer_num = 400
gen_lamda = []
t = 0

hr = 0
ct_index = 1

def GeometricPoissonDistribution(n, theta, lamda):
    if n == 0:
        return math.exp(-lamda)
    else:
        res = 0
        for k in list(range(1 , n + 1)):
            value = math.exp(-lamda) * ((lamda ** k) / math.factorial(k)) * ((1 - theta) ** (n - k)) * (theta ** k)
            tem = math.factorial(n - 1,)/(math.factorial(k - 1)*math.factorial(n - k ))
            value = value*tem
            #math.comb(n - 1, k - 1))
            res += value
        return res

def DistancePoolGen(mean, var, min, max, size = 10000, distribution = 'normal'):
    """
    주어진 분포에 따라 거리 pool을 반환
    :param mean: normal시 평균
    :param var: normal시 분산
    :param min: uniform min
    :param max: uniform max
    :param size: 거리 분포 pool의 크기
    :param distribution: 거리 분포 타입 normal / uniform
    :return:
    """
    if distribution == 'normal':
        ct_dist = np.random.normal(mean, var, size = size)
    elif distribution == 'uniform':
        ct_dist = np.random.uniform(min, max, size=size)
    else:
        print('Wong input')
        pass
    num = 0
    dist_pool = []
    for dist in ct_dist:
        if dist < min or dist > max:
            pass
        else:
            dist_pool.append(dist)
            num += 1
    while num < size:
        val = np.random.choice(ct_dist)
        if val < min or val > max:
            pass
        else:
            dist_pool.append(val)
            num += 1
    return dist_pool

def RadiusPoolGen(size = 10000):
    """
    무작위 각도 pool을 반환
    :param size: pool 크기
    :return: 무작위 각도 pool
    """
    anglepool = []
    num = 0
    while num < size*2:
        anglepool.append(2 * math.pi * np.random.random())
        num += 1
    return anglepool

def LocGen(dist1, dist2, angle1, angle2, center = [25,25]):
    """
    입력 된 데이터에 따라 2개의 point를 생성
    :param dist1: 중심으로 부터 가게 사이의 거리
    :param dist2: 중심으로 부터 고객 사이의 거리
    :param angle1: 중심과 가게의 각도
    :param angle2:  중심과 고객의 각도
    :param center: 중심 좌표
    :return: [가게 x, 가게 y] , [고객 x, 고객 y]
    """
    o_x = dist1 * math.cos(angle1) + center[0]
    o_y = dist1 * math.sin(angle1) + center[1]
    d_x = dist2 * math.cos(angle2) + o_x
    d_y = dist2 * math.sin(angle2) + o_y
    return [o_x, o_y], [d_x, d_y]


def LogGenRandom(dist1, angle, max_x, max_y):
    o_x = random.randrange(1, max_x - 1)
    o_y = random.randrange(1, max_y - 1)
    d_x = dist1 * math.cos(angle) + o_x
    d_y = dist1 * math.sin(angle) + o_y
    return [o_x, o_y], [d_x, d_y]


def NextMin(lamda):
    """
    다음에 고객이 발생할 시점을 예측
    :return: 다음 사건 발생 시점
    """
    # lambda should be float type.
    # lambda = input rate per unit time
    next_min = (-math.log(1.0 - np.random.random()) / lamda)*60
    return float("%.2f" % next_min)

def ExcelSaver(infos, header=['Num','x1','y1','x2','y2','dist','far','interval'], save_name = 'test'):
    file_name = save_name + 'loc_data.xlsx'
    wb = Workbook()
    wb['Sheet'].title = 'loc'
    sheet_sm = wb['loc']
    row = 1
    for col in range(1, len(header) + 1):
        sheet_sm.cell(row=row, column=col).value = header[col-1]
    row += 1
    for info in infos:
        col = 1
        for ele in info:
            sheet_sm.cell(row=row, column=col).value = ele
            col += 1
        row += 1
    wb.save(filename=file_name)


def ReadCSV(dir, start = 1):
    res = []
    with open(dir, 'r') as raw:
        reader = csv.reader(raw)
        index = 0
        for lines in reader:
            if index >= start:
                res.append([float(lines[1]), float(lines[2])])
            index += 1
    return res


Points = []
dist1pool = DistancePoolGen(15, 2, 0, 40) # 중심과 일반 고객 사이의 거리
dist2pool = DistancePoolGen(3, 1, 0, 7, distribution='uniform') #중심과 가게 사이의 거리
angle1pool = RadiusPoolGen()
angle2pool = RadiusPoolGen()
distlongpool = DistancePoolGen(30, 5, 0, 40) # 거리가 먼 고객의 거리를 생성

LongRatio = 0.2
dist_pools = []
num = 1
cor_x = []
cor_y = []
lamda_multiplier = 4 # 이 계수를 조정해 하루에 생성되는 고객의 수를 조절
lamda = [2,5,9,7,6,5,7,11,15,16,13,10,8.5,5.2,2.5] #합 : 122
rev_lamda = []

for i in lamda:
    rev_lamda.append(int(lamda_multiplier*i))
print(rev_lamda)
env_time = 0
"""
while num < 1000 + 1:
    rv = np.random.random()
    long_para = 0
    dist1 = np.random.choice(dist1pool)
    dist2 = np.random.choice(dist2pool)
    distLong = np.random.choice(distlongpool)
    angle1 = np.random.choice(angle1pool)
    angle2 = np.random.choice(angle2pool)
    if rv > LongRatio:
        #store, ct = LocGen(dist2, dist1, angle2, angle1)
        store, ct = LocGen(dist2, dist1, angle2, angle1, center = store)
    else:
        #store, ct = LocGen(dist2, distLong, angle2, angle1)
        store, ct = LocGen(dist2, dist1, angle2, angle1, center=store)
        long_para = 1
    time_slot = int(env_time//60)
    if time_slot >= len(rev_lamda) - 1: #종료 시점에 종료.
        print('Time break', time_slot)
        break
    interval = NextMin(rev_lamda[time_slot])
    dist = ((store[0] - ct[0])**2 + (store[1] - ct[1])**2)**(1/2)
    gen = [num] + store + ct + [dist] + [long_para] + [interval]
    #print(gen)
    Points.append(gen)
    dist_pools.append(dist)
    cor_x.append(store[0])
    cor_x.append(ct[0])
    cor_y.append(store[1])
    cor_y.append(ct[1])
    num += 1
    env_time += interval
"""
randomratio = 0.3
max_x = 50
max_y = 50

dir_random = '데이터/Random_loc_data_1.csv'
dir_cluster = '데이터/MaternCluster loc_data_2.csv'
randompool = ReadCSV(dir_random)
clusterpool = ReadCSV(dir_cluster)

while num < 1000 + 1:
    rv = np.random.random()
    dist1 = np.random.choice(dist1pool)
    angle1 = np.random.choice(angle1pool)
    range_para = False
    while range_para == False:
        #store, ct = LogGenRandom(dist1, angle1, max_x, max_y)
        #랜덤 인스턴스
        store = None
        if rv < 0.3:
            store = random.choice(randompool)
        else:
            store = random.choice(clusterpool)
        d_x = dist1 * math.cos(angle1) + store[0]
        d_y = dist1 * math.sin(angle1) + store[1]
        ct = [d_x, d_y]
        dist = dist = ((store[0] - ct[0])**2 + (store[1] - ct[1])**2)**(1/2)
        if 0 < ct[0] < max_x and 0 < ct[1] < max_y:
            range_para = True
            break
    time_slot = int(env_time//60)
    if time_slot >= len(rev_lamda) - 1: #종료 시점에 종료.
        print('Time break', time_slot)
        break
    interval = NextMin(rev_lamda[time_slot])
    dist = ((store[0] - ct[0])**2 + (store[1] - ct[1])**2)**(1/2)
    gen = [num] + store + ct + [dist] + [0] + [interval]
    #print(gen)
    Points.append(gen)
    dist_pools.append(dist)
    cor_x.append(store[0])
    cor_x.append(ct[0])
    cor_y.append(store[1])
    cor_y.append(ct[1])
    num += 1
    env_time += interval

#생성된 점들에 대하여 50x50 크기를 벗어나는 점들을 반영.
min_x = int(abs(min(cor_x))) + 1
min_y = int(abs(min(cor_y))) + 1
print('minx', min_x, 'miny', min_y)
rev_Points = [[0, 25+min_x, 25 + min_y , 25+min_x, 25+min_y,0,0,0]]
for info in Points:
    data = [info[0],info[1] + min_x,info[2] + min_y, info[3] + min_x,info[4] + min_y]
    data += info[5:]
    rev_Points.append(data)

ExcelSaver(rev_Points, save_name = 'test_0131_random_cluster_v1')
print('Dist Mean', sum(dist_pools)/len(dist_pools))


"""
### 라이더 lamda 생성 과정 ####
interval_lamda_base =[2,2,5,4,7,4,3,2,2,4,7,7,3,3,3]
interval_lamda = []
for i in interval_lamda_base:
    interval_lamda.append(i*4)

interval_data = []
#Interval Maker
for ite in range(100):
    dr_t = 0
    next_time = 0
    tem = []
    while dr_t <= 900:
        next_time = NextMin(interval_lamda[int(dr_t // 60)])
        tem.append(next_time)
        dr_t += next_time
    tem2 = [ite]
    for info in tem:
        if type(info) == float:
            tem2.append(info)
    interval_data.append(tem2)

print('lamda len',len(interval_data))
"""
"""
file_name = 'interval_rider_data.csv'
wb = Workbook()
wb['Sheet'].title = 'loc'
sheet_sm = wb['loc']
row = 1
for info in interval_data:
    print(info[-5:])
    col = 1
    for ele in info:
        sheet_sm.cell(row=row, column=col).value = ele
        col += 1
    row += 1

wb.save(filename=file_name)
"""
"""
csvfile = open('interval_rider_data4.csv','w', newline="")

csvwriter = csv.writer(csvfile)
for row in interval_data:
    csvwriter.writerow(row)
csvfile.close()
"""