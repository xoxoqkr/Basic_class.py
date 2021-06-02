# -*- coding: utf-8 -*-
import random
import simpy
import math
import operator
import numpy as np
from openpyxl import Workbook
from datetime import datetime
#from sklearn.cluster import KMeans
from ASP_code import AssignPSolver
import copy
from ASP_code import subsidyASP
from itertools import combinations



class Customer(object):
    def __init__(self, env, name, input_location=[[25, 25], [25, 25]], end_time = 60, ready_time=3, service_time=3, fee = 1000, select_pr = [0.5,0.5]):
        self.name = name  # 각 고객에게 unique한 이름을 부여할 수 있어야 함. dict의 key와 같이
        self.time_info = [round(env.now, 2), None, None, None, None, end_time, ready_time, service_time]
        # [0 :발생시간, 1: 차량에 할당 시간, 2:차량에 실린 시간, 3:목적지 도착 시간,
        # 4:고객이 받은 시간, 5: 보장 배송 시간, 6:가게에서 준비시간,7: 고객에게 서비스 하는 시간]
        self.gen_time = int(env.now)
        self.location = input_location
        self.assigned = False
        self.loaded = False
        self.done = False
        self.cancelled = False
        self.server_info = []
        self.fee = fee
        self.fees = []
        self.paid_fee = None
        self.selecting_time = []
        self.doneby = None
        #self.pr = select_pr
        pl = self.PlatformChoice(select_pr)
        self.pr_res = pl
        self.selected_t = None
        #self.pr_res = random.choices(list(range(len(self.pr))), weights= self.pr, k = 1) #가지고 있는 선택 함수로, 주문할 플랫폼을 결정.
        env.process(self.Decline(env))


    def PlatformChoice(self, pr):
        pl_names = list(range(len(pr)))
        selected = random.choices(pl_names, weights= pr, k = 1)
        return selected[0]

    def Decline(self, env):
        yield env.timeout(self.time_info[5])
        if self.assigned == False and self.done == False:
            self.cancelled = True
            self.server_info.append(["N", env.now])


class Driver(object):
    def __init__(self, env, name, speed, customer_set, end_time=800, thres = 30, max_on_hand_order = 3, wageForHr = 10000):
        self.env = env
        self.name = name
        self.route = []#[[0, 1, [25, 25]]]
        # [[이름,유형,위치],..,] 유형 = {0:o_loc, 1:d_loc, 2:idle_loc}
        self.veh = simpy.Resource(env, capacity=1)
        self.last_location = [25, 25]  # 송파 집중국
        self.idle_location = [25, 25]
        self.serving = []
        self.served = [[0, 1, [25, 25]]] #[0] # [[이름,유형,위치],..,] 유형 = {0:o_loc, 1:d_loc, 2:idle_loc}
        self.speed = speed
        self.get_more = True
        self.left = False
        self.route_check = [[0,1,[25,25],0]] # [[이름,유형,위치,시간],..,] 유형 = {0:o_loc, 1:d_loc, 2:idle_loc}
        self.due_times = []
        self.thres = thres
        self.max_on_hand_order = max_on_hand_order
        self.wageForHr = wageForHr
        self.idle_times = []
        self.minimum_wage = 0
        self.select_infos = []
        self.earn_fee = []
        self.gen_time = int(env.now)
        self.left_time = None
        env.process(self.Runner(env, customer_set))
        env.process(self.CustomerSelector(env, customer_set, max_on_hand_order = self.max_on_hand_order))
        env.process(self.DriverLeft(env))


    def DriverLeft(self, env):
        """
        일정 조건을 만족하는 기사의 경우에는 시장에서 이탈.
        :param env:
        """
        yield env.timeout(120)
        self.left = True
        self.left_time = int(env.now)


    def GetMore(self, policy_max = 3, num_on_hand = False):
        names = []
        for info in self.route:
            if info[0] not in names and info[0] > 0:
                names.append(info[0])
        if num_on_hand == True:
            return int(len(names))
        if len(names) < policy_max:
            return True
        else:
            return False


    def CustomerSelector(self, env, customer_set, wait_time=3, end_time=800, try_num=2, max_on_hand_order = 3):
        """
        고객 중 가장 높은 가치를 가지는 고객을 선택하는 함수.
        매 interval 마다 고객을 탐색하게 된다.
        env:
        customer_set:
        wait_time: 검색을 시도하는 interval time
        end_time: 운행 종료 시점
        try_num: 단일 고객 검색 시, 라이더가 고객을 검색해 보는 횟수
        :rtype: None :: 고객을 선택하는 작업 수행
        """
        yield env.timeout(5)
        while env.now <= end_time and self.left == False:
            get_more = self.GetMore(policy_max = max_on_hand_order)
            #print("GET MORE",get_more, self.route)
            if get_more == True: #현재 가지고 있는 주문의 수가 get_more보다 더 적은 경우 -> 고객을 추가로 선택할 수 있음
                for ite in range(0,try_num): # 희망 고객 설정 여러번 도전할 수 있음.
                    ava_cts = UnloadedCustomer(customer_set, env.now)
                    if len(ava_cts) > 0:
                        priority_orders = PriorityOrdering(self, ava_cts, int(env.now))
                        if len(priority_orders) > 0:
                            #print("Value Compare", priority_orders[0][1] , self.minimum_wage)
                            self.select_infos.append([round(env.now,1), priority_orders[:3], customer_set[priority_orders[0][0]].fees])
                            if priority_orders[0][1] > self.minimum_wage:
                                insert_infos = priority_orders[0]
                                customer = customer_set[insert_infos[0]]
                                #print('insert_infos',insert_infos)
                                #todo : 삽입될 순서를 다시 계산 해야 함. 만약 고객이 삽입되게 된다면, 삽입 정보가 바뀌기 때문.
                                #feasibility, route = CheckThisCT2(self, customer, insert_infos[2], customer_set, env.now, thres= self.thres)
                                feasibility, route = CheckThisCT3(self, customer, insert_infos[2], customer_set, env.now, thres= self.thres)
                                if feasibility == True:
                                    if insert_infos[4] > customer.fee:
                                        print('TIME',env.now ,'rider', self.name, 'select',customer.name,'E_value:',insert_infos[4], customer.fee,'subsidy paid')
                                    else:
                                        print('TIME', env.now, 'rider', self.name, 'select', customer.name, 'E_value:',
                                              insert_infos[4],customer.fee, 'None subsidy')
                                    tem = []
                                    for info in priority_orders:
                                        tem.append([info[0], round(info[1],1)])
                                    print('others',tem)
                                    if customer.pr_res == 1: #todo : 보조금을 지불하는 고객만.
                                        customer.paid_fee = insert_infos[3]
                                    customer.assigned = True
                                    self.route = route
                                    self.serving.append(customer.name)
                                    customer.selected_t = env.now
                                    self.earn_fee.append(customer.fee + MySubsidy(self, customer, int(env.now)))
                                    self.minimum_wage = round(max(self.minimum_wage * 1.1,1000), 2)
                                    # 라우트에 추가
                            else: # 고객의 이윤이 낮으면, 자신의 가치를 할인해서 수행.
                                self.minimum_wage = round(min(self.minimum_wage*0.9,0),2)
                                pass
                    else:
                        pass  # 추가로 고른 고객은 서비스 품질 상 맞지 않음.
                yield env.timeout(wait_time)
            else: #이미 고객을 get_more보다 많이 가지고 있는 경우 -> wait time만큼 기다렸다가 다시 고객 탐색.
                yield env.timeout(wait_time)
                # 정책에 의해 추가 고객 선택이 불가.

    def Runner(self, env, customer_set, end_time=800, wait_time=3):
        """
        라이더가 고객을 선택하면, 고객을 서비스 하도록 수행하는 과정을 표현
        :param env:
        :param customer_set:
        :param end_time:
        :param wait_time:
        """
        #init_num = self.name + 1
        #self.route.append([init_num,0,customer_set[init_num].location[0]])
        #self.route.append([init_num, 1, customer_set[init_num].location[1]])
        #customer_set[init_num].assigned = True
        #env.process(self.Driving(env, customer_set[1],customer_set, loc_type=2))
        while env.now <= end_time:
            if len(self.route) > 0:
                #print('Now', env.now)
                #print("self route", self.route)
                ct_name = self.route[0][0]
                customer = customer_set[ct_name]
                print(env.now ,'Go ct', customer.name, self.route[0])
                #print('appened info',self.route[0])
                yield env.process(self.Driving(env, customer,customer_set, loc_type = self.route[0][1]))
                #print(self.route[0][1], "done", env.now)
                added = copy.deepcopy(self.route[0])
                self.served.append(added)
                self.served[-1].append(env.now)
                del self.route[0]
                #print("test2")
            else:
                # 고객을 고르려는 행위 넣기
                yield env.timeout(wait_time)  # 이 시간 만큼 대기
                #print("Rider", self.name, 'waits from', env.now - wait_time, '~', env.now)

    def Driving(self, env, customer, customer_set ,loc_type = 2):
        """
        실제로 경로를 움직이는 함수
        :param env:
        :param customer:
        :param loc_type:
        """
        customer.assigned = True
        with self.veh.request() as req:
            req.info = [customer.name, loc_type]
            yield req  # users에 들어간 이후에 작동
            if customer.name > 0:
                bf_loc = customer_set[self.served[-1][0]].location[self.served[-1][1]]
                #bf_loc = self.veh.users[0].info[2]
            else:
                bf_loc = [30,30]
            if loc_type == 0: #위치가 주문의 가게인 경우
                customer.time_info[1] = env.now
                time = distance(bf_loc, customer.location[loc_type]) / self.speed
                req.info.append(customer.location[loc_type])
                self.due_times.append([customer.name, loc_type, env.now, env.now + time + customer.time_info[6]])
                yield env.timeout(time)
                yield env.timeout(customer.time_info[6])
                #self.serving.append(customer.name)
                customer.time_info[2] = env.now
                customer.loaded = True
                customer.server_info.append([self.name, 0, env.now])
                self.route_check.append([customer.name, loc_type, customer.location[loc_type] , env.now])
                print(round(env.now,2),'Veh#', self.name, 'arrived store for CT#',customer.name,'`s store After route:',self.route)
            elif loc_type == 1: #위치가 주문의 도착점(고객의 집)인 경우
                #print('points2', customer.name, 'customer0', env.now, self.route)
                time = distance(bf_loc, customer.location[loc_type]) / self.speed
                req.info.append(customer.location[loc_type])
                self.due_times.append([customer.name, loc_type, env.now, env.now + time + customer.time_info[7]])
                yield env.timeout(time)
                customer.time_info[3] = env.now
                yield env.timeout(customer.time_info[7])
                customer.done = True
                customer.time_info[4] = env.now
                customer.server_info.append([self.name, 1,env.now])
                self.route_check.append([customer.name, loc_type, customer.location[loc_type],env.now])
                self.serving.remove(customer.name)
                #self.served.append([customer.name, env.now])
                print(round(env.now,2),'Veh#', self.name, 'arrived CT#', customer.name,'dep. at' ,customer.time_info[2],'After route:',self.route)
            else: #고객 주문이 아닌 지점이 경로에 있는 경우 ->에러의 소지가 있음.
                req.info.append(self.idle_location)
                time = distance(bf_loc, self.idle_location) / self.speed
                yield env.timeout(time)
                #print('points3', customer.name, '??', env.now, self.route)
            #print('req.info', req.info)
            #print(self.name, "CT:", customer.name, loc_type, customer.location, "done", env.now)

    def NextCustomerSearch(self):
        """
        자기가 가진 주문의 수에 따라, 다음 주문 선택 시점이 달라짐.
        :return: 다음 주문을 선택하려는 시간
        """
        time_list = [1,2,3,4] #주문이 0,1,2,3개 일때, 차량이 다음 주문을 선택하려는 시간.
        my_order_num = self.GetMore(num_on_hand = True)
        return time_list[my_order_num]


def NextMin(lamda):
    """

    :rtype: object
    """
    # lambda should be float type.
    # lambda = input rate per unit time
    next_min = (-math.log(1.0 - random.random()) / lamda)*60
    return float("%.2f" % next_min)


def CheckThisCT3(veh, customer, insert_infos, customer_set, now_time, thres = 20):
    """
    차량에 고객이 할당 되었을 때, 해당 고객의 삽입이, (1)삽입된 고객의 TW를 지킬 수 있는지
    (2)기존 고객들의 TW를 지킬 수 있는지 여부를 계산.
    :param veh: class veh
    :param customer: class customer
    :param insert_infos: [o 삽입 위치, d 삽입위치]
    :param thres: 분  : 고객의 음식이 라이더에게 실린 후 배송 되어야 하는 시간
    :return: True/False , 갱신된 라우트 경로
    """
    #print('checkthis CT')
    route = copy.deepcopy(veh.route)
    index = 0
    for info in insert_infos:
        route.insert(info, [customer.name, index, customer.location[index]])
        index += 1
    current_serving2 = {}
    for ct_info in route:
        if ct_info[0] > 0 and ct_info[0] not in current_serving2:
            current_serving2[ct_info[0]] = [None, None]
    #print("test route", route)
    time = 0
    if len(veh.route) > 0:
        for info in veh.due_times:
            if veh.route[0][0] == info[0] and veh.route[0][1] == info[1]:
                time = time + (info[3] - now_time)
                #print('add time 0 - >', time,'::', veh.route[0], info, '::',now_time)
    for node_index in range(0, len(route)):
        if route[node_index][0] == 0:
            pass
        elif node_index == 0:
            if route[node_index][1] == 0:
                current_serving2[route[node_index][0]][0] = time
            else:
                current_serving2[route[node_index][0]][1] = time
        else:
            bf = route[node_index - 1][2]
            af = route[node_index][2]
            time += (distance(bf, af) / veh.speed)
            ct_name = route[node_index][0]
            if route[node_index][1] == 0: #todo :  각 노드에서 걸리는 시간을 계산 해야 함. 가게와 고객을 서비스하는데 걸리는 시간.
                time += customer_set[ct_name].time_info[6]
            else:
                time += customer_set[ct_name].time_info[7]
            loc_type = route[node_index][1]
            current_serving2[route[node_index][0]][loc_type] = time
    for ct_name in current_serving2:
        if current_serving2[ct_name][0] == None:
            for info in veh.route_check:
                if ct_name == info[0] and info[1] == 0:
                    current_serving2[ct_name][0] = info[3] - now_time
        elif current_serving2[ct_name][1] == None:
            print('Ref Error')
        else:
            pass
    #print('route',route)
    #print('current_serving2',current_serving2)
    for ct_name in current_serving2:#(1)해당 고객 O-D시간이 일정 이내, (2)기존 고객들의 O-D시간이 위반 하는지 계산.
        if current_serving2[ct_name][1] -  current_serving2[ct_name][0] > thres\
                or current_serving2[ct_name][1] + now_time - customer_set[ct_name].time_info[0] > customer.time_info[5]:
            return False, None
    return True, route

def PriorityOrdering(veh, ava_customers, env_time):
    """
    veh의 입장에서 ava_customers를 가치가 높은 순서대로 정렬한 값을 반환
    :param veh: class veh
    :param ava_customer_names: 삽입 가능한 고객들의 class의 list
    :return: [[고객 이름, 증가하는 거리, [o index, d index]],...] 이미 오름 차순으로 정렬됨.
    """
    res = []
    for customer in ava_customers:
        #print('PPP', customer.name, customer.location)
        info, cost = CalScore(veh, customer)
        cost = Won(cost, veh.speed, wageForHr = veh.wageForHr)
        fee = MySubsidy(veh.name, customer, env_time)
        value = CalValue(fee - cost, simple_diff= True)
        #print(fee, cost, value)
        if fee > cost:
            #print('size check',fee, cost)
            pass
        res.append([customer.name, value , info, fee - customer.fee, fee, cost])
    res.sort(key=operator.itemgetter(1), reverse = True)
    return res


def MySubsidy(veh_name, customer,now_time):
    """
    고객을 배송했을 떄, 차량이 받을 수 있는 보조금을 계산
    :param veh_name:
    :param customer:
    :return: 라이더가 받는 보조금
    """
    index_list = list(range(len(customer.fees)))
    index_list.reverse()
    if customer.pr_res == 1: #todo : 현재는 보조금을 주는 것과 안주는 것만 존재.
        for info_index in index_list:
            info = customer.fees[info_index]
            if veh_name == info[0] and info[2] >= now_time - 20:
                res = info[1]
                return customer.fee + res
            elif info[0] == 'all': #모두에게 주는 보조금인 경우
                res = info[1]
                return customer.fee + res
            else:
                pass
        return customer.fee
    else:
        return customer.fee


def CalScore(veh, customer):
    """
    입력 된 고객이 차량 경로의 어디에 삽입되면 좋을 지를 계산.
    경로의 끝에도 삽입될 수 있음.
    기존 경로 + O-D
    :param veh: class veh
    :param customer: 삽입되고자 하는 고객 class
    :return: [o 삽입 위치, d 삽입 위치], 비용
    """
    rev_route = copy.deepcopy(veh.route)
    org_res = RouteInsertCost(rev_route, 1, customer, 0)
    if len(org_res) > 0:
        #print("insert cost1", org_res)
        org_insert = org_res[0][0]
        rev_route.insert(org_insert, [customer.name, 0, customer.location[0]])
        d_start_index = org_insert + 1 # 가게 위치의 삽입으로 그 보다 하나 더 뒤부터 고객은 위치할 수 있음.
        des_res = RouteInsertCost(rev_route, d_start_index, customer,1)
        des_insert = des_res[0][0]
        rev_route.insert(des_insert, [customer.name, 1, customer.location[1]])
        #print("rev route2", rev_route)
        return [org_insert, des_insert], org_res[0][1] + des_res[0][1]
    else:
        return [0, 1], distance(customer.location[0],customer.location[1])


def CalValue(val, enchor = None, simple_diff = False):
    """
    주어진 값에 대한 가치함수 값을 계산
    :param val:
    :param enchor:
    :param simple_diff:
    :return: 가치함수 값
    """
    if val < 0:
        return 0
    elif simple_diff == True:
        return val
    val = min(10, val)
    if enchor == None:
        return math.exp(val)
    elif val > enchor:
        return math.exp(val - enchor)
    else:
        return 0


def HowMore(veh, customer, target, inc = 1,  error_thres = 10):
    """
    현재 매력도 보다 큰 값인 target 매력도가 되기 위해 증가시켜야 하는 보조금을 계산
    :param veh:
    :param customer:
    :param target: 목표로 하는 매력의 값
    :param inc: 한번 연산시 증가하는 보조금의 양
    :param error_thres: 연산 종료 error 오차
    :return: target 매력이 되기 위해 지급해야 보조금의 양
    """
    #print('howmore', veh,customer)
    info, cost = CalScore(veh,customer)
    cost = Won(cost, veh.speed, wageForHr = veh.wageForHr) #todo : 해당 업무를 하는데 필요한 시간(시간단위)*시급
    add = inc
    base = customer.fee - cost
    value = min(10, base)
    ite = 0
    #print('How more start')
    while value - target < error_thres and ite < 100:
        value = CalValue(value + add, simple_diff= True)
        add += inc
        ite += 1
        #print(value, add, ite)
    print('How more end')
    return add


def RouteInsertCost(route, start_index, customer, loc_type):
    """
    해당 고객의 loc_type이 차량에 삽입되는데 필요한 비용(추가 이동 거리)을 계산
    :param route:
    :param start_index:
    :param customer:
    :param loc_type:
    :return:
    """
    res = []
    for insert_index in range(start_index, len(route) + 1):
        bf = route[insert_index - 1][2]
        if insert_index < len(route): #좌 우에 노드가 있는 위치에 삽입되는 경우.
            af = route[insert_index][2]
            org_cost = distance(bf, af)
            rev_cost = distance(bf, customer.location[loc_type]) + distance(customer.location[loc_type], af)
        else: #마지막 자리에 삽입된다면, 원래 비용은 0 이고, 그냥 추가 삽입 비용이 생기는 것.
            org_cost = 0
            rev_cost = distance(bf, customer.location[loc_type])
        res.append([insert_index, rev_cost - org_cost])
    res.sort(key=operator.itemgetter(1))
    return res

def distance(x1, x2):
    return round(math.sqrt((x1[0] - x2[0]) ** 2 + (x1[1] - x2[1]) ** 2), 2)
"""

def distance(x1, x2, dist_cal_type = 'euc'):
    if dist_cal_type == 'euc':
        return round(math.sqrt((x1[0] - x2[0]) ** 2 + (x1[1] - x2[1]) ** 2), 2)
    else:
        return round(abs(x1[0] - x2[0]) + abs(x1[1] - x2[1]), 2)

"""



def UnloadedCustomer(customer_set, now_time, pl_num = None):
    """
    아직 라이더에게 할당되지 않은 고객들을 반환
    :param customer_set:
    :param now_time: 현재시간
    :return: [고객 class, ...,]
    """
    res = []
    if pl_num != None:
        for pl_index in range(pl_num):
            res.append([])
    for ct_name in customer_set:
        customer = customer_set[ct_name]
        cond1 = now_time - customer.time_info[0] < customer.time_info[5]
        cond2 = customer.assigned == False and customer.loaded == False and customer.done == False
        cond3 = customer.time_info[1] == None
        if cond1 == True and cond2 == True and cond3 == True and ct_name > 0 and len(customer.server_info) == 0:
            if pl_num == None:
                res.append(customer)
            else:
                res[customer.pr_res].append(customer)
    return res

def AvaRider(rider_set, max_num = 3):
    """
    현재 고객 주문을 수행할 수 있는(현재 주문 수가 GetMore를 만족하는) 라이더들을 계산

    :param rider_set:
    :param max_num:
    :return:
    """
    riders = []
    for rider_index in rider_set:
        rider = rider_set[rider_index]
        if rider.GetMore(policy_max = max_num) == True and rider.left == False:
            riders.append(rider.name)
        else:
            pass
    print("AvaRider",riders)
    return riders

def Won(dist, speed,wageForHr = 10000):
    return ((dist/speed)/60)*wageForHr

def ProblemInput(rider_set, customer_set, now_time, pl_num = None, thres = 30):
    riders = AvaRider(rider_set)
    cts = UnloadedCustomer(customer_set, now_time, pl_num = pl_num)
    values = []
    d_orders = []
    for rider_name in riders:
        rider = rider_set[rider_name]
        d_orders.append([rider_name, len(rider.route)])
        for customer in cts:
            #print('ProblemInput', rider, customer)
            loc, cost = CalScore(rider, customer)
            cost = Won(cost, rider.speed, wageForHr=rider.wageForHr)
            value = customer.fee - cost
            if value > 0:
                values.append(CalValue(value,simple_diff= True))
            else:
                values.append(0)
    v_old = np.array(values).reshape(len(riders), len(cts))
    cts_name = []
    for ct in cts:
        cts_name.append(ct.name)
    d_orders.sort(key = operator.itemgetter(1))
    d_orders_res = [] #todo : 현재는 라우트의 길이가 짧을 수록 고객을 더 먼저 선택할 것이라고 가정함.
    for rider_name in riders:
        index = 1
        for info in d_orders:
            if rider_name == info[0]:
                d_orders_res.append(index)
                break
            index += 1
    ava_match = []
    for rider_name in riders:
        rider = rider_set[rider_name]
        for customer in cts:
            ava = Ava_rider_ct(rider, customer, customer_set, now_time, thres = thres)
            ava_match.append(ava)
    ava_match = np.array(ava_match).reshape(len(riders), len(cts))
    return v_old, riders, cts_name, d_orders_res, ava_match

def AvaArrayMaker(rider_names, cts_names, riders, customer_set, urgent_cts, now_time):
    """
    Matrix{라이더, 고객} 크기의 매트릭스
    1인 경우 라이더가 해당 고객을 서비스할 수 있음(고객의 TW를 지키면서)
    0인 경우 라이더가 해당 고객을 서비스 할 수 없음.(해당 차량이 고객을 서비스 하는 경우 고객의 TW를 지킬 수 없음)
    :param rider_names:
    :param cts_names:
    :param riders:
    :param customer_set:
    :param urgent_cts:
    :param now_time:
    :return:
    """
    res = []
    for rider_name in rider_names:
        rider = riders[rider_name]
        for ct in cts_names:
            customer = customer_set[ct]
            feasibility, res = Ava_rider_ct(rider, customer, customer_set, now_time)
            if feasibility == False:
                res.append(0)
            else:
                res.append(1)
    res = np.array(res).reshape(len(rider_names), len(cts_names))
    return res

def Ava_rider_ct(veh, customer, customer_set, now_time, thres = 20):
    """
    이 고객이 해당 차량에 삽입될 수 있는지를 계산
    :param veh:
    :param customer:
    :param customer_set:
    :param now_time:
    :param thres: 고객의 TW를 넘어갈 수 있는 시간
    :return: 1(삽입 가능)/ 0(삽입 불가능)
    """
    ava_indexs = list(combinations(list(range(len(veh.route))), 2))
    for indexs in ava_indexs:
        feasibility, res = CheckThisCT3(veh, customer, indexs, customer_set, now_time, thres = thres)
        if feasibility == False:
            return 0
    return 1


def FeeUpdater2(rev_v, customer_set, riders, rider_set ,cts_name, now_time):
    """
    구한 보조금을 고객에게 반영
    고객들의 fees를 update.
    :param rev_v:
    :param customer_set:
    :param riders:
    :param rider_set:
    :param cts_name:
    :param now_time:
    :return:
    """
    print('Time', now_time)
    for info in rev_v:
        rider_name = riders[info[0]]
        rider = rider_set[rider_name]
        customer = customer_set[cts_name[info[1]]]
        inc = HowMore(rider, customer, info[2], inc=50)
        #customer.fees.append([rider_name, inc, now_time]) #todo : 보조금 지급이 가장 앞으로 옮.
        customer.fees = [[rider_name, inc, now_time]] #todo: 보조금 지급을 갱신함
        #customer.fees = [['all', inc, now_time]] #todo: 모든 라이더가 선택할 수 있는 보조금
        print('CT#',customer.name,'rider',customer.fees[-1][0],'subsidy',customer.fees[-1][1], customer.fee)
    return True


def solver_interpreterForIP2(res_x, res_v, v_old):
    """
    IP가 푼 문제의 해를 판단
    :param res_x:
    :param res_v:
    :param v_old:
    :return:
    """
    x = AssignPSolver.indexReturner2D(res_x, 1)
    v = AssignPSolver.indexReturner2DBiggerThan0(res_v, 1)
    if len(v) > 0:
        print('res_v')
        print(res_v)
        print(res_v + v_old)
        return True, v
    else:
        return False, None

def WhoGetPriority(cts_names , customer_set, ava_riders_num, now_time, time_thres = 0.7):
    """
    현재 TW가 임박한 고객을 계산
    :param cts_names:
    :param customer_set:
    :param ava_riders_num:가능한 라이더의 수(이 수만큼만 고객들을 뽑기 때문)
    :param now_time: 현재 시간
    :param time_thres: 고객의 endtime 중 얼마의 시간이 지난 경우 우선 고객으로 지정하는가? (0<x<1)
    :return: [고객 이름1, 고객 이름2,...ㅡ]
    """
    scroes = []
    for ct_name in cts_names:
        customer = customer_set[ct_name]
        if now_time - customer.time_info[0] > customer.time_info[5]*time_thres:
            scroes.append([customer.name, customer.time_info[0]])
    if len(scroes) > 0:
        scroes.sort(key=operator.itemgetter(1))
        res = []
        tem = []
        for info in scroes[:ava_riders_num]:
            res.append(cts_names.index(info[0]))
            tem.append(info[0])
        return res, tem
    return [], []


def OneClickSolver(driver_names, customer_names, v_old, rider_set ,customer_set ,rider_seq = None, lower = None, upper = None, ava_match = None ,priority = None,now_time = 0, print_gurobi = False, C_p_para = 1):
    """

    :param driver_names: 가능한 라이더 이름들
    :param customer_names: 가능한 고객 이름들
    :param v_old:
    :param rider_set: 모든 라이더 class
    :param customer_set: 모든 고객 class
    :param rider_seq: 가능한 라이더들의 선택 순서
    :param lower:
    :param upper:
    :param ava_match:
    :param priority:
    :param now_time:
    :param print_gurobi:
    :return:
    """
    if rider_seq == None:
        res, vars = subsidyASP.SimpleInverseSolver(driver_names, customer_names, v_old, print_gurobi = print_gurobi, sp = priority, lower_b = lower)
    else:
        if C_p_para == 1:
            old_x, old_obj = subsidyASP.SimpleAssingmentSolver(driver_names, customer_names, v_old)
            print('old_obj', old_obj)
            # input('obj check')
            effec = []
            for row in range(0,len(old_x)):
                for col in range(0,len(old_x[0])):
                    if old_x[row,col] == 1:
                        effec.append(col)
            print('effec cts',effec)
            #res = subsidyASP.SimpleInverseSolverwithX(driver_names, customer_names, old_x, v_old, old_obj)
            res, vars = subsidyASP.SimpleInverseSolver3(driver_names, customer_names, v_old, rider_seq,
                                                        ava_match=ava_match,
                                                        print_gurobi=print_gurobi, sp=effec, minimum_fee=lower)
        else:
            res, vars = subsidyASP.SimpleInverseSolver3(driver_names, customer_names, v_old, rider_seq, ava_match = ava_match,
                                                        print_gurobi=print_gurobi, sp=priority, minimum_fee = lower)
    feasibility, res2 = solver_interpreterForIP2(res[0], res[1], v_old)
    if feasibility == True:
        print('FeeUpdater')
        #print(res2)
        #input('Enter')
        #FeeUpdater(res2, customer_set, driver_names, rider_set ,customer_names, now_time)
        FeeUpdater2(res2, customer_set, driver_names, rider_set, customer_names, now_time)
        print('FeeUpdater end')
        return res2
    return feasibility

def TwoPhaseSolver(rider_set, customers_set,driver_names, rider_seq,customer_names,v_old, priority_orders, now_time):
    """
    Phase1에서 vo를 계산 (vo : 라이더의 선택을 바꾸기 위해서 필요한 보조금의 하한 개념)
    Phase2에서 vo를 기반으로 하는 계산 수행.
    :param rider_set: 모든 차량  class
    :param customers_set: 모든 고객 class
    :param driver_names: 가능한 차량 이름들
    :param rider_seq: 가능한 차량의 순서
    :param customer_names: 가능한 고객 이름들
    :param v_old:
    :param priority_orders: 우선 순위 고객들
    :param now_time:
    :return:
    """
    vo = subsidyASP.InversePhase1(rider_set, rider_seq, customers_set)
    #print('driver`s num' ,len(driver_names),driver_names)
    #print('minimum subsidy len',len(vo), vo)
    #print('v_old size',np.shape(v_old),v_old)
    res, vars = subsidyASP.InversePhase2(driver_names, customer_names, v_old, vo,sp=priority_orders)
    feasibility, res2 = solver_interpreterForIP2(res[0], res[1], v_old)
    if feasibility == True:
        FeeUpdater2(res2, customers_set, driver_names, rider_set, customer_names, now_time)
        print('FeeUpdater end')
        return res2
    return feasibility

def DriverMaker(env, driver_dict, customer_set ,speed = 2, end_time = 800, min_interval = 10,max_interval = 30, thres = 30, max_on_hand_order = 3, intervals = [], interval_res = [],interval_para = False):
    name = 0
    while env.now < end_time:
        rider = Driver(env, name, speed, customer_set, max_on_hand_order = max_on_hand_order, thres= thres)
        driver_dict[name] = rider
        if interval_para == False:
            #next_time = random.randrange(min_interval, max_interval)
            #intervals.append(next_time)
            print('Hr',intervals[int(env.now//60)])
            next_time = NextMin(intervals[int(env.now//60)])
            #print(next_time)
            #input('Stop')
            interval_res.append(next_time)
        else:
            next_time = intervals[name]
        name += 1
        yield env.timeout(next_time)

def storeAndcustomer(max_cor, thres, speed = 2):
    """
    1. 가게는 0~max_cor안에서 무작위로 발생.
    2. 가게에서 부터 thres분 이내에 도착할 수 있는 범위에서 고객의 위치를 발생시킴.
    :param max_cor:
    :param thres:
    :param speed:
    :return:
    """
    store = [random.randrange(0, max_cor),random.randrange(0, max_cor)]
    while True:
        customer = [random.randrange(0, max_cor),random.randrange(0, max_cor)]
        if distance(store, customer)/speed < thres:
            return [store, customer]

def CustomerMaker(env, customer_dict,  end, init , start = 1, end_time = 800, max_cor = 50, max_interval = 5, fee = 1500, speed = 2, dist_thres = 20):
    for name in range(start, init + 1):
        #cor = [[random.randrange(0, max_cor),random.randrange(0, max_cor)],
        #       [random.randrange(0, max_cor),random.randrange(0, max_cor)]]
        cor = storeAndcustomer(max_cor, dist_thres, speed =  speed)
        customer = Customer(env, name, input_location = cor)
        customer_dict[name] = customer
    name = len(customer_dict)
    while env.now < end_time and name < end:
        #cor = [[random.randrange(0, max_cor), random.randrange(0, max_cor)],
        #       [random.randrange(0, max_cor), random.randrange(0, max_cor)]]
        cor = storeAndcustomer(max_cor, dist_thres, speed =  speed)
        customer = Customer(env, name, input_location = cor, fee = fee)
        customer_dict[name] = customer
        name += 1
        next_time = random.randrange(1,max_interval)
        yield env.timeout(next_time)



def PlatformRunner(env, riders, customers, end_time = 800,problemType = 4, run_interval = 5, print_gurobi = False, Problem_states = None, lower = 500, pl_num = None, para2 = True, thres = 30,
                   C_p_para = 1, C_p_time_thres = 0.7):
    """

    :param env:
    :param riders:
    :param customers:
    :param end_time:
    :param problemType:
    :param run_interval:
    :param print_gurobi:
    :param Problem_states:
    :param lower:
    :param pl_num:
    :param para2: True :원 모형 풀이 / False = 변형 모형 풀이
    """
    while env.now < end_time:
        v_old, rider_names, cts_name, d_orders_res, ava_match = ProblemInput(riders, customers, env.now, pl_num = pl_num, thres = thres)
        print('Now',env.now,"Env info/CT#:", len(cts_name) ,':AvaRider#:',len(rider_names))
        #input('Type Enter 2')
        urgent_cts, tem1 = WhoGetPriority(cts_name, customers, len(rider_names), env.now, time_thres= C_p_time_thres)
        print('urgent_cts',urgent_cts)
        if len(cts_name) >= len(rider_names) >= 2 and len(urgent_cts) > 0:
            #urgent_cts, tem1 = WhoGetPriority(cts_name, customers, len(rider_names), env.now)
            tem = []
            for ct in urgent_cts:
                tem.append(customers[ct])
            #print('urgent_cts')
            #print(urgent_cts, tem1)
            p_v = []
            for row in v_old:
                for v in row:
                    p_v.append(int(v))
            p_v = np.array(p_v).reshape(len(rider_names),len(cts_name))
            print(p_v)
            print('ava match',ava_match)
            if para2 == True:
                v_info = OneClickSolver(rider_names, cts_name,v_old, riders, customers, ava_match = [],
                                        rider_seq = d_orders_res, now_time= env.now, print_gurobi = print_gurobi, priority = urgent_cts, lower = lower,
                                        C_p_para = C_p_para)
            else:
                v_info = TwoPhaseSolver(riders, customers,rider_names, d_orders_res,cts_name,v_old, urgent_cts, env.now)
            if Problem_states != None:
                Problem_states.append([env.now, len(cts_name), len(rider_names), len(urgent_cts), v_info])
        yield env.timeout(run_interval)
        print('Now',env.now,"Env info End")
        #input('check')

def DataSaveAsXlxs(veh_set, customer_set, problem_states, veh_write=False, ct_write=False, add_infos=None, thres1 = 90, thres2 = 20):
    now = datetime.today().isoformat()
    file_name = 'IP' + now[5:7] + "-" + now[8:10] + "-" + now[11:13] + "-" + now[14:16] + 'res.xlsx'
    if add_infos != None:
        file_name = str(add_infos) + file_name
    print("file names", file_name)
    wb = Workbook()
    wb['Sheet'].title = 'summary'
    sheet_sm = wb['summary']
    header_summary = ['Served #', 'ave ltd', 'ave food ltd', 'ltd violate#', 'ave violate time','bin ltd violate#', 'ave bin violate time','doneby']
    for col in range(1, len(header_summary) + 1):
        sheet_sm.cell(row=1, column=col).value = header_summary[col - 1]
    res_ltd = []
    res_f_ltd = []
    violate1 = []
    violate2 = []
    for ct_name in customer_set:
        info = customer_set[ct_name].time_info
        if info[4] != None:
            res_ltd.append(info[4] - info[0])
            res_f_ltd.append(info[4]- info[2])
            if info[4] - info[0] >thres1:
                violate1.append(info[4] - info[0] - thres1)
            if info[4]- info[2] > thres2:
                violate2.append(info[4] - info[2] - thres2)
    summary_res = [len(res_ltd), sum(res_ltd)/max(1,len(res_ltd)), sum(res_f_ltd)/max(1,len(res_f_ltd)),
                   len(violate1),sum(violate1)/max(1,len(violate1)),len(violate2),sum(violate2)/max(1,len(violate2))]
    for col in range(1, len(summary_res) + 1):
        sheet_sm.cell(row=2, column=col).value = summary_res[col - 1]

    row = 5
    header_problem_states = ['Time', 'cts#', 'ridr#', 'urgent cts#','v info']
    for col in range(1, len(header_problem_states) + 1):
        sheet_sm.cell(row=row, column=col).value = header_problem_states[col - 1]
    row += 1
    for info in range(0, len(problem_states)):
        for col in range(1, len(problem_states[info]) + 1):
            if col < 5:
                sheet_sm.cell(row=row, column=col).value = problem_states[info][col - 1]
            else:
                sheet_sm.cell(row=row, column=col).value = str(problem_states[info][col - 1])
        row += 1
    if veh_write == True:
        # 차량 정보 입력
        # 평균 거리, 유휴 시간, 평균 서비스한 인원 수
        sheet_veh = wb.create_sheet('veh')
        header_veh = ['veh #', 'Served_ct_#', "dist per ct", 'served', 'select infos']
        for col in range(1, len(header_veh) + 1):
            sheet_veh.cell(row=1, column=col).value = header_veh[col - 1]
        row_num = 2
        for veh_name in veh_set:
            lt = []
            veh = veh_set[veh_name]
            for ct_info in veh.served:
                ct_name = ct_info[0]
                if customer_set[ct_name].done == True and customer_set[ct_name].time_info[4] != None:
                    lt.append(customer_set[ct_name].time_info[4] - customer_set[ct_name].time_info[0])
            route = []
            for info in veh.served:
                route.append(info[2])
            route_time, cut_index = RouteTime(route, customer_set, cor_parar = True)
            infos = [veh.name, len(lt), route_time/ max(len(lt),1), str(veh.served), str(veh.select_infos)]
            for col in range(1, len(infos) + 1):
                sheet_veh.cell(row=row_num, column=col).value = infos[col - 1]
            row_num += 1
    if ct_write == True:
        sheet_ct = wb.create_sheet('ct')
        header_ct = ['No.', 'Occurred', 'store','Served','food_lt', 'fee','subsidy', 'cor_x', 'cor_y','server info', 'doneby','라이더가 선택한시간']
        # [발생시간, 차량에 할당 시간, 차량에 실린 시간, 목적지 도착 시간, 고객이 받은 시간]
        for col in range(1, len(header_ct) + 1):
            sheet_ct.cell(row=1, column=col).value = header_ct[col - 1]
        for key in customer_set:
            info = [key]
            ct = customer_set[key]
            for index in [0,2,4]:
                info.append(ct.time_info[index])
            if ct.time_info[4] != None:
                try:
                    info.append(info[3] - info[2])
                except:
                    info.append(None)
            else:
                info.append(0)
            info.append(ct.fee)
            info.append(ct.paid_fee)
            info.append(str(customer_set[key].location[0]))
            info.append(str(customer_set[key].location[1]))
            info.append(str(customer_set[key].server_info))
            info.append(customer_set[ct_name].pr_res )
            info.append(customer_set[key].selected_t)
            ele_count = 1
            start_row = sheet_ct.max_row + 1
            for ele in info:
                sheet_ct.cell(row=start_row, column=ele_count).value = ele
                ele_count += 1
    wb.save(filename=file_name)

def CustomerGeneratorForIP(env, customer_list, dir=None, end_time=10000, speed = 2, wagePerHr = 10000, select_pr = [0.5 , 0.5], customer_wait_time = 40):
    datas = open(dir, 'r')
    lines = datas.readlines()
    for line in lines[2:]:
        data = line.split(';')
        store_loc = [float(data[1]), float(data[2])]
        customer_loc = [float(data[3]), float(data[4])]
        fee = distance(store_loc, customer_loc)*200 + 2500 #기본 수수료에 거리마다 100원의 추가 요금이 발생하는 요금제
        #fee = Won(distance(store_loc, customer_loc), speed, wagePerHr) + 1500
        c = Customer(env, int(data[0]), input_location = [store_loc, customer_loc], fee= fee, select_pr = select_pr, end_time= customer_wait_time)
        print('this customer select mode',c.pr_res)
        customer_list[int(data[0])] = c
        print('Time',round(env.now,2) ,'CT#', c.name, 'gen')
        yield env.timeout(float(data[5]))
        if env.now > end_time:
            break


def RouteTime(route, customer_set, max_time=120, cor_parar = False, speed = 2):
    res = []
    cut_index = len(route)
    cut_index_deter = False
    for index in list(range(1, len(route))):
        # print("index", index, route[index - 1],route[index])
        if cor_parar == False:
            bf = customer_set[route[index - 1]].location
            af = customer_set[route[index]].location
        else:
            bf = route[index - 1]
            af = route[index]
        res.append(distance(bf, af))
        if sum(res)/speed > max_time and cut_index_deter == False:
            cut_index = index
            cut_index_deter = True
    # print("route time check", route[:5], round(sum(res),2), cut_index)
    return round(sum(res), 2), cut_index


def indexReturner1D(list1D, val):
    res = []
    for index in range(len(list1D)):
        if int(list1D[index]) == val:
            res.append(index)
    return res

def indexReturner2D(list2D, val):
    res = []
    for row in range(len(list2D)):
        for index in range(len(list2D[row])):
            if int(list2D[row][index]) == val:
                res.append([row, index])
    return res


def DayByDayStore(customer_set, old_store_prs, re_ratio, day, time_thres = 90, revise_rule = 1):
    """
    1. 이전 날의 결과에 따라 다음날에 어떤 플랫폼에 주문을 할당할지를 결정.
    선택 확률의 재분배는 아래와 같이 계산됨.
    pr = 전날 i플랫폼의 선택확률*(1 - 재분배율) + 재분배율 * (i플랫폼의 성공 주문 수 /전체 성공 주문 수)
    :param customer_set: 이전 날의 고객 obj
    :param num_p: 문제 상황의 플랫폼 수
    :param old_store_prs: 이전 날의 주문 선택확률
    :param re_ratio: 재분배율
    :param revise_rule: 어떤 방식으로 갱신하는가 1:평균 리드타임에 반비례 2:평균 서비스한 고객 수에 비례 3:성공률(각자 플랫폼이 받은 고객 수 대비 성공한 고객의 비율)
    :return: 다음날의 선호 확률
    """
    print('daybyday', len(customer_set))
    sucess = [0] * len(old_store_prs)
    ltd = [0] * len(old_store_prs)
    paid_fees = [0] * len(old_store_prs)
    paid_subsidy = [0] * len(old_store_prs)
    paid_num = [0] * len(old_store_prs)
    #for i in range(len(old_store_prs)):
    #    sucess.append([])
    org_num = [0] * len(old_store_prs)
    for customer_name in customer_set:
        customer = customer_set[customer_name]
        org_num[customer.pr_res] += 1
        print(customer.name, customer)
        if customer.done == True and customer.time_info[4] - customer.time_info[0] <= time_thres:
            print(customer.name, customer.done, customer.time_info[4] - customer.time_info[0], customer.time_info[4],
                  '-', customer.time_info[0])
            assigend_platform = customer.pr_res
            sucess[assigend_platform] += 1
            ltd[assigend_platform] += (customer.time_info[4] - customer.time_info[0])
            paid_fees[assigend_platform] += customer.fee
            if customer.paid_fee != None:
                paid_subsidy[assigend_platform] += customer.paid_fee
                if customer.paid_fee > 1:
                    paid_num[assigend_platform] += 1
    ave_ltd = [0] * len(old_store_prs)
    ave_paid_fee = [0] * len(old_store_prs)
    ave_paid_subsidy = [0] * len(old_store_prs)
    for pl in range(len(old_store_prs)):
        ave_ltd[pl] = round(ltd[pl]/sucess[pl],4)
        ave_paid_fee[pl] = round(paid_fees[pl]/sucess[pl],4)
        ave_paid_subsidy[pl] = round(paid_subsidy[pl]/sucess[pl],4)
    store_prs = [0] * len(old_store_prs)
    if revise_rule == 1:
        # 평균 리드타임에 반비례하여, 다음 날의 플랫폼 선택 확률이 변하는 경우
        pr = [0,0]
        pr[1] = ave_ltd[0]/(ave_ltd[0]+ave_ltd[1])
        pr[0] = ave_ltd[1] / (ave_ltd[0] + ave_ltd[1])
        store_prs[0] = old_store_prs[0]*(1-re_ratio) +re_ratio*pr[0]
        store_prs[1] = old_store_prs[1]*(1-re_ratio) +re_ratio*pr[1]
    elif revise_rule == 2:
        # 평균 서비스한 고객 수에 비례하여, 다음 날의 플랫폼 선택 확률이 변하는 경우
        print('success',sucess)
        sucess_sum = sum(sucess)
        for index in range(len(old_store_prs)):
            pr = old_store_prs[index]*(1-re_ratio) + re_ratio*(sucess[index]/sucess_sum)
            store_prs[index] = pr
    else:
        #성공률(각자 플랫폼이 받은 고객 수 대비 성공한 고객의 비율)
        pl1 = sucess[0] / org_num[0]
        pl2 = sucess[1] / org_num[1]
        pls = [pl1, pl2]
        for index in range(len(old_store_prs)):
            pr = old_store_prs[index] * (1 - re_ratio) + re_ratio*(pls[index]/sum(pls))
            store_prs[index] = pr
    res = []
    for i in range(len(old_store_prs)):
        tem = [day, i, old_store_prs[i] , sucess[i], ave_ltd[i], ave_paid_fee[i], ave_paid_subsidy[i], paid_num[i]]
        res.append(tem)
    return store_prs, res

def daysDataSave(infos, add_infos = None):
    now = datetime.today().isoformat()
    file_name = 'Days_info_IP' + now[5:7] + "-" + now[8:10] + "-" + now[11:13] + "-" + now[14:16] + 'res.xlsx'
    if add_infos != None:
        file_name = str(add_infos) + file_name
    print("file names", file_name)
    wb = Workbook()
    wb['Sheet'].title = 'summary'
    sheet_sm = wb['summary']
    sheet_sm.cell(row=1, column=1).value = add_infos
    header_summary = ['day', 'pl num', 'selection pr', 'served #', 'ave ltd','ave paid fee', 'ave paid subsidy','subsidy paid ct#']
    row = 2
    for col in range(1, len(header_summary) + 1):
        sheet_sm.cell(row=row, column=col).value = header_summary[col - 1]
    row += 1
    for info in infos:
        for col in range(1, len(info) + 1):
            sheet_sm.cell(row= row, column=col).value = info[col - 1]
        row += 1
    wb.save(filename=file_name)


def DataSaver2(scenario_info, customer_set, driver_set ,thres, speed, fee_ratio = 0.2, now_time = 800):
    infos = [str(scenario_info)]
    total_paid_subsidy = []
    total_paid_fees = []
    subsidy_paid_ct_num = 0
    #고객 대기 시간 계산
    ltd_ave = []
    f_ltd_ave = []
    ltds = [] #
    f_ltds = [] #음식 대기 시간
    ltd_val = [0,20,40,60,80,100,120,140,160,180]
    f_val = [0,5,10,15,20,25,30,35]
    ct_paid_fee = 0
    total_dist = 0
    for _ in ltd_val[1:]:
        ltds.append([])
    for _ in f_val[1:]:
        f_ltds.append([])
    for ct_name in customer_set:
        ct = customer_set[ct_name]
        if ct.done == True:
            total_dist += distance(ct.location[0], ct.location[1])
            ltd = round(ct.time_info[4] - ct.time_info[0],2)
            f_ltd = round(ct.time_info[3] - ct.time_info[2],2)
            ltd_ave.append(ltd)
            f_ltd_ave.append(f_ltd)
            for index in range(0,len(ltd_val)-1):
                if ltd_val[index] <= ltd < ltd_val[index + 1]:
                    ltds[index].append(ltd)
                    break
            for index in range(0,len(f_val)-1):
                if f_val[index] <= f_ltd < f_val[index + 1]:
                    f_ltds[index].append(f_ltd)
                    break
            ct_paid_fee += ct.fee
            total_paid_subsidy.append(ct.paid_fee)
            if ct.paid_fee > 0:
                subsidy_paid_ct_num += 1
    infos += [len(ltd_ave), np.mean(ltd_ave), np.std(ltd_ave)]
    for t in ltds:
        infos.append(len(t))
    infos.append('//')
    infos += [len(f_ltd_ave), np.mean(f_ltd_ave), np.std(f_ltd_ave)]
    for t in f_ltds:
        infos.append(len(t))
    infos += [thres, speed]
    infos.append('//')
    infos.append(int(sum(total_paid_subsidy)))
    infos.append(int(sum(total_paid_subsidy)/len(total_paid_subsidy)))
    infos.append(subsidy_paid_ct_num)
    working_time = []
    for driver_name in driver_set:
        driver = driver_set[driver_name]
        total_paid_fees.append(sum(driver.earn_fee))
        if driver.left_time == None:
            working_time.append(int(now_time - driver.gen_time)) #아직 운행 중
        else:
            working_time.append(int(driver.left_time - driver.gen_time)) #떠남
    infos.append('//')
    infos.append(fee_ratio)
    infos.append(ct_paid_fee*fee_ratio - int(sum(total_paid_subsidy)))
    rider_total_earn = ct_paid_fee * (1 - fee_ratio) + int(sum(total_paid_subsidy))
    infos.append(rider_total_earn)
    infos.append(rider_total_earn/len(total_paid_fees))
    infos.append(rider_total_earn /(sum(working_time)/60))
    infos.append(len(total_paid_fees))
    infos.append(ct_paid_fee * fee_ratio - int(sum(total_paid_subsidy)))
    infos.append('//')
    infos.append(total_dist/len(ltd_ave))
    """
    infos.append(int(sum(total_paid_fees)*fee_ratio))
    infos.append(int(sum(total_paid_fees) * (1 - fee_ratio)))
    infos.append(int(sum(total_paid_fees)*(1 - fee_ratio) / len(total_paid_fees)))
    infos.append(int((sum(total_paid_fees) * (1 - fee_ratio))/(sum(working_time)/60)))
    infos.append(len(total_paid_fees))
    infos.append(int(sum(total_paid_fees) * fee_ratio) - int(sum(total_paid_subsidy)))
    """
    return infos

def DataSaver2_summary(infos):
    header = ['시나리오명','서비스된 고객수','평균','표준편차','0-20','20-40','40-60','60-80'
              ,'80-100','100-120','120-140','140-160','160~','//'
              ,'음식리드타임','평균','표준변차','0~5','5~10','10~15','15~20','20~25','25~30','30~'
              ,'thres','라이더 속도','//'
              ,'지급된 보조금','평균보조금','보조금 지급 받은 주문 건수','//'
              ,'플랫폼 수수료율','플랫폼 수익','라이더 수익','라이더 평균 수익','라이더 시간당 이익','발생 라이더수','플랫폼 순수익(플랫폼 수익-지급된 보조금)']
    now = datetime.today().isoformat()
    file_name = 'scenario_compete' + now[5:7] + "-" + now[8:10] + "-" + now[11:13] + "-" + now[14:16] + 'res.xlsx'
    wb = Workbook()
    wb['Sheet'].title = 'summary'
    sheet_sm = wb['summary']
    col = 1
    for row in range(1, len(header) + 1):
        sheet_sm.cell(row=row, column=col).value = header[row-1]
    col = 2
    for info in infos:
        row = 1
        for ele in info:
            sheet_sm.cell(row=row, column=col).value = ele
            row += 1
        col += 1
    wb.save(filename=file_name)

def DataSaver3_summary(infos):
    header = ['시나리오명','서비스된 고객수','평균','표준편차','0-20','20-40','40-60','60-80'
              ,'80-100','100-120','120-140','140-160','160~','//'
              ,'음식리드타임','평균','표준변차','0~5','5~10','10~15','15~20','20~25','25~30','30~'
              ,'thres','라이더 속도','//'
              ,'지급된 보조금','평균보조금','보조금 지급 받은 주문 건수','//'
              ,'플랫폼 수수료율','플랫폼 수익','라이더 수익','라이더 평균 수익','라이더 시간당 이익','발생 라이더수','플랫폼 순수익(플랫폼 수익-지급된 보조금)','//'
              ,'배송 고객 평균 거리','라이더 lamda곱']
    now = datetime.today().isoformat()
    file_name = 'ITE_scenario_compete' + now[5:7] + "-" + now[8:10] + "-" + now[11:13] + "-" + now[14:16] + 'res.xlsx'
    wb = Workbook()
    wb['Sheet'].title = 'summary'
    sheet_sm = wb['summary']
    col = 1
    for row in range(1, len(header) + 1):
        sheet_sm.cell(row=row, column=col).value = header[row-1]
    col = 2
    print('infos check')
    print(infos)
    master_infos = []
    for sc_infos in infos:
        if sc_infos != []:
            sc_info1 = [sc_infos[0][0]]
            for val_index in range(1,len(sc_infos[0])):
                tem = []
                if sc_infos[0][val_index] == '//':
                    sc_info1.append('//')
                else:
                    for info in sc_infos:
                        #print('info raw')
                        #print(info)
                        tem.append(info[val_index])
                    #print(tem)
                    sc_info1.append(round(sum(tem)/len(tem),2))
            master_infos.append(sc_info1)
    for info in master_infos:
        row = 1
        for ele in info:
            sheet_sm.cell(row=row, column=col).value = ele
            row += 1
        col += 1
    wb.save(filename=file_name)

def RandomSubsidy(env, driver_set, customer_set, cp_t = 30,max_s = 0.15, run_interval = 10, end_time = 600):
    """
    cp_t만큼 지난 고객들에게, random.randrange(0, max_s, int(max_s/10)) 만큼의 보조금을 할당
    할당된 보조금은 모든 라이더가 접근할 수 있음.
    :param env:
    :param driver_set: 라이더 셋 [라이더 class, ..., ]
    :param customer_set: 고객 셋 [고객 class, ...,]
    :param cp_t: 최대 할당 보조금
    :param max_s: 기존 customer.fee 대비 지급하려는 보조금 비율
    :param run_interval: 보조금 지급 간격
    """
    yield env.timeout(60) # warm up time
    while env.now < end_time:
        #1 보조금 필요한 상황 인식
        #2 보조금 할당
        if len(driver_set) > 2:
            ct_num = random.randrange(1,len(driver_set))
            ava_cts = UnloadedCustomer(customer_set, env.now)
            if len(ava_cts) > ct_num:
                cts = random.sample(ava_cts, ct_num) #cts = [고객 class, 고객 class,...,]
                now = int(env.now)
                cts_names = []
                for ct in cts:
                    cts_names.append([ct.name, now - ct.time_info[0]])
                cts_names.sort(key=operator.itemgetter(1))
                cts_urgent = []
                for ct_info in cts_names:
                    if ct_info[1] >= cp_t:
                        cts_urgent.append(customer_set[ct_info[0]])
                # 3 보조금 갱신
                for ct in cts_urgent:
                    customer = customer_set[ct.name]
                    fee = int(customer.fee*max_s)
                    subsidy = random.randrange(0, fee, int(fee/10))
                    customer.fees = [['all', subsidy]] #일괄적인 보조금
                    #customer.fees.append(['all', subsidy])
        yield env.timeout(run_interval)


def AllSubsidy(env, driver_set, customer_set, inc_ratio_step = [[30,50,0.05],[50,70,0.1],[70,100,0.15]], run_interval = 10, end_time = 600):
    """
    일정 시간이 지난 고객에게 일괄적으로 보조금을 할당하는 문제.
    paid_step은 지난 시간 list [30,60,80] 생성 후 30분 지난 뒤, 생성 후 60분 지난 뒤, 생성 후 80분 지난뒤 각각
    customer.fee*inc_ratio^ 만큼의 보조금을 더 지급
    :param env:
    :param driver_set: 라이더 셋 [라이더 class, ..., ]
    :param customer_set: 고객 셋 [고객 class, ...,]
    :param inc_ratio_step: [[발생후 시간, 원래 금액 대비 더해지는 보조금 %],...,]
    :param run_interval: 보조금 지급 간격
    """
    yield env.timeout(60)  # warm up time
    while env.now < end_time:
        #1 보조금 필요 상황 인식
        #2 보조금 할당
        ava_cts = UnloadedCustomer(customer_set, env.now) #ava_cts = [고객 class, 고객 class,...,]
        # 3 보조금 갱신
        now = int(env.now)
        for ct in ava_cts:
            customer = customer_set[ct.name]
            for info in inc_ratio_step:
                wait_time = now - customer.time_info[0]
                if info[0] <= wait_time < info[1]:
                    subsidy = int(customer.fee * info[2])
                    customer.fees = [['all', subsidy]]  # 일괄적인 보조금
                    #customer.fees.append(['all', subsidy])
                    break
        yield env.timeout(run_interval)
