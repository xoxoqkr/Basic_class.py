# -*- coding: utf-8 -*-
import  numpy as np
import Basic_class as Basic
from openpyxl import Workbook
from datetime import datetime
import math

def SingleDataSaver(scenario_info, customer_set, driver_set ,thres, speed, now_time, fee_ratio = 0.2):
    infos = [str(scenario_info)]
    total_paid_subsidy = []
    total_paid_fees = []
    subsidy_paid_ct_num = 0
    ltd_ave = []
    f_ltd_ave = []
    ltds = [] #
    f_ltds = [] #음식 대기 시간
    ltd_val = [0,20,40,60,80,100,120,140,160,180]
    f_val = [0,5,10,15,20,25,30,35]
    ct_paid_fee = 0
    total_dist = 0
    far_num = 0
    assigned_time = []
    fees = []
    for _ in ltd_val[1:]:
        ltds.append([])
    for _ in f_val[1:]:
        f_ltds.append([])
    for ct_name in customer_set:
        ct = customer_set[ct_name]
        if ct.done == True and ct.name > 0:
            total_dist += Basic.distance(ct.location[0], ct.location[1])
            ltd = round(ct.time_info[4] - ct.time_info[0],2)
            f_ltd = round(ct.time_info[3] - ct.time_info[2],2)
            assigned_time.append(round(ct.time_info[1] - ct.time_info[0],2))
            ltd_ave.append(ltd)
            f_ltd_ave.append(f_ltd)
            for index in range(0,len(ltd_val)-1):
                if ltd_val[index] <= ltd < ltd_val[index + 1]:
                    ltds[index].append(ltd)
                    break
            for index in range(0,len(f_val)-1):
                if f_val[index] <= f_ltd < f_val[index + 1]:
                    f_ltds[index].append(f_ltd)
                    break
            ct_paid_fee += (ct.fee[0] + ct.fee[1])
            total_paid_subsidy.append(ct.fee[1])
            if ct.fee[1] > 0:
                subsidy_paid_ct_num += 1
                fees.append(ct.fee[1])
            if ct.far == 1:
                far_num += 1
    infos += [len(ltd_ave), np.mean(ltd_ave), np.std(ltd_ave)]
    for t in ltds:
        infos.append(len(t))
    infos.append('//')
    infos += [len(f_ltd_ave), np.mean(f_ltd_ave), np.std(f_ltd_ave)]
    for t in f_ltds:
        infos.append(len(t))
    infos += [thres, speed]
    infos.append('//')
    infos.append(int(sum(total_paid_subsidy)))
    infos.append(int(sum(total_paid_subsidy)/len(total_paid_subsidy)))
    infos.append(subsidy_paid_ct_num)
    working_time = []
    idle_t1 = []
    idle_t2 = []
    fee_analyzer = []
    subsidy_analyzer = []
    for slot_num in range(int(math.ceil(now_time / 60))):
        fee_analyzer.append([])
        subsidy_analyzer.append([])
    #fee_analyzer = []*(16) #[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
    #subsidy_analyzer = []*(16) #[[], [], [], [], [], [], [], [], [], [], [], [], [], [], []]
    for driver_name in driver_set:
        driver = driver_set[driver_name]
        total_paid_fees.append(sum(driver.earn_fee))
        if driver.left_time == None:
            working_time.append(int(now_time - driver.gen_time)) #아직 운행 중
        else:
            working_time.append(int(driver.left_time - driver.gen_time)) #떠남
        idle_t1.append(sum(driver.idle_times[0]))
        idle_t2.append(sum(driver.idle_times[1]))
        time_slot = 0
        for info in driver.fee_analyze:
            if len(info) > 0:
                fee_analyzer[time_slot].append(sum(info))
            time_slot += 1
        time_slot = 0
        for info in driver.subsidy_analyze:
            if len(info) > 0:
                subsidy_analyzer[time_slot].append(sum(info))
            time_slot += 1
    idle_t1 = round(sum(idle_t1)/len(idle_t1),2)
    idle_t2 = round(sum(idle_t2) / len(idle_t2), 2)
    infos.append('//')
    infos.append(fee_ratio)
    infos.append(ct_paid_fee*fee_ratio - int(sum(total_paid_subsidy)))
    rider_total_earn = ct_paid_fee * (1 - fee_ratio) + int(sum(total_paid_subsidy))
    infos.append(rider_total_earn)
    infos.append(rider_total_earn/len(total_paid_fees))
    infos.append(rider_total_earn /(sum(working_time)/60))
    infos.append(len(total_paid_fees))
    infos.append(ct_paid_fee * fee_ratio - int(sum(total_paid_subsidy)))
    infos.append('//')
    infos.append(total_dist/len(ltd_ave))
    infos.append('//')
    gen_time = list(range(0,15))
    tem = []
    for _ in gen_time[1:]:
        tem.append([])
    for ct_name in customer_set:
        ct = customer_set[ct_name]
        if ct.done == True and ct.fee[1] > 0:
            for index in range(0,len(gen_time)-1):
                if gen_time[index] <= ct.time_info[1]//60 < gen_time[index + 1]:
                    tem[index].append(ct.time_info[1])
                    break
    for t in tem:
        infos.append(len(t))
    infos.append(0)
    infos.append('//')
    gen_nums = []
    served_num = []
    for _ in gen_time[1:]:
        gen_nums.append(0)
        served_num.append(0)
    tem = []
    for _ in gen_time[1:]:
        tem.append([])
    for ct_name in customer_set:
        ct = customer_set[ct_name]
        gen_nums[int(ct.time_info[0]//60)] += 1
        if ct.done == True:
            served_num[int(ct.time_info[0] // 60)] += 1
    for num in gen_nums:
        infos.append(num)
    infos.append(0)
    infos.append('//')
    for num in served_num:
        infos.append(num)
    infos.append(0)
    infos.append('//')
    infos.append('N/A')
    infos.append(far_num)
    infos.append(sum(assigned_time)/len(assigned_time))
    infos.append(idle_t1)
    infos.append(idle_t2)
    infos.append('//')
    for info in fee_analyzer:
        infos.append(sum(info))
    infos.append('//')
    for info in subsidy_analyzer:
        infos.append(sum(info))
    infos.append('//')
    try:
        infos.append(round(np.mean(fees),2))
        infos.append(round(np.std(fees),2))
        infos.append(max(fees))
        infos.append(min(fees))
        infos.append('//')
    except:
        infos += ['N/A','N/A','N/A','N/A','//']
    return infos


def DataSaver4_summary(infos, saved_name = 'None'):
    header = ['시나리오명','서비스된 고객수','평균','표준편차','0-20','20-40','40-60','60-80'
              ,'80-100','100-120','120-140','140-160','160~','//'
              ,'서비스 받은 고객 수','평균','표준변차','0~5','5~10','10~15','15~20','20~25','25~30','30~'
              ,'thres','라이더 속도','//'
              ,'지급된 보조금','평균보조금','보조금 지급 받은 주문 건수','//'
              ,'플랫폼 수수료율','플랫폼 수익','라이더 수익','라이더 평균 수익','라이더 시간당 이익','발생 라이더수','플랫폼 순수익(플랫폼 수익-지급된 보조금)','//'
              ,'배송 고객 평균 거리','//','보조금 받은 고객 수0-1','1-2','2-3','3-4'
              ,'4-5','5-6','6-7','7-8','8-9','9-10','10-11','11-12','12-13','13-14','14-15','//'
              ,'생성0-1','1-2','2-3','3-4','4-5','5-6','6-7','7-8','8-9','9-10','10-11','11-12','12-13','13-14','14-15','//'
              ,'서비스0-1','1-2','2-3','3-4','4-5','5-6','6-7','7-8','8-9','9-10','10-11','11-12','12-13','13-14','14-15','//'
              ,'None','원거리 서비스 고객 수','기사 선택까지 시간','유휴시간0 평균','유휴시간1 평균','//'
              ,'시간대별 기본금0-1', '1-2', '2-3', '3-4','4-5', '5-6', '6-7', '7-8', '8-9', '9-10', '10-11', '11-12', '12-13', '13-14', '14-15','//'
              ,'시간대별 보조금0-1', '1-2', '2-3', '3-4', '4-5', '5-6', '6-7', '7-8', '8-9', '9-10', '10-11', '11-12', '12-13', '13-14', '14-15', '//'
            ,'mean', 'std', 'max', 'min','//' #여기 까지가 single save에 저장되는 것.
              ,'제안된 보조금 수','제안된 보조금 수0-1', '1-2', '2-3', '3-4', '4-5', '5-6', '6-7', '7-8', '8-9', '9-10', '10-11', '11-12', '12-13','13-14', '14-15', '//']
    now = datetime.today().isoformat()
    file_name = saved_name + now[5:7] + "-" + now[8:10] + "-" + now[11:13] + "-" + now[14:16] + 'res.xlsx'
    wb = Workbook()
    wb['Sheet'].title = 'summary'
    sheet_sm = wb['summary']
    col = 1
    for row in range(1, len(header) + 1):
        sheet_sm.cell(row=row, column=col).value = header[row-1]
    col = 2
    print('infos check')
    print(infos)
    master_infos = []
    for sc_infos in infos:
        if sc_infos != []:
            sc_info1 = [sc_infos[0][0]]
            for val_index in range(1,len(sc_infos[0])):
                tem = []
                if sc_infos[0][val_index] == '//':
                    sc_info1.append('//')
                else:
                    for info in sc_infos:
                        tem.append(info[val_index])
                    try:
                        sc_info1.append(round(sum(tem)/len(tem),2))
                    except:
                        sc_info1.append('None')
            master_infos.append(sc_info1)
    for info in master_infos:
        row = 1
        for ele in info:
            sheet_sm.cell(row=row, column=col).value = ele
            row += 1
        col += 1
    wb.save(filename=file_name)
    print(file_name + ' save done')

def CustomerDataSaver(customer_set, saved_name):
    now = datetime.today().isoformat()
    file_name = saved_name + now[5:7] + "-" + now[8:10] + "-" + now[11:13] + "-" + now[14:16] + 'res.xlsx'
    wb = Workbook()
    wb['Sheet'].title = 'customer'
    sheet_sm = wb['customer']
    header = ['name', 'gen_t', 'assigned_t', 'served_t','doneBy','dist']
    row = 1
    for col in range(1, len(header) + 1):
        sheet_sm.cell(row=row, column=col).value = header[col-1]
    ct_datas = []
    for ct_name in customer_set:
        ct = customer_set[ct_name]
        if ct.done == True:
            ct_datas.append([ct.name, ct.time_info[0], ct.time_info[1], ct.time_info[4],ct.server_info[0], Basic.distance(ct.location[0], ct.location[1])])
        else:
            ct_datas.append([ct.name, ct.time_info[0], 'N/A', 'N/A','N/A', Basic.distance(ct.location[0], ct.location[1])])
    row = 2
    for info in ct_datas:
        col = 1
        for ele in info:
            sheet_sm.cell(row=row, column=col).value = ele
            col += 1
        row += 1
    wb.save(filename=file_name)
    print(file_name + ' save done')


def DataSave(data, RIDER_DICT, CUSTOMER_DICT, insert_thres, speed, run_time,subsidy_offer,subsidy_offer_count,ite,mean_list,std_list, add_info = ''):
    fees = []
    subsidy_take_num = []
    for ct_name in CUSTOMER_DICT:
        customer = CUSTOMER_DICT[ct_name]
        fees.append(customer.fee[0])
        if customer.server_info != None and customer.server_info[0] == customer.fee[2]:
            subsidy_take_num.append([customer.name, customer.server_info[0], customer.fee[1]])
    #print('Ave fee', sum(fees) / len(fees))
    #print('subsidy_take_num', len(subsidy_take_num))
    info = SingleDataSaver(data, CUSTOMER_DICT, RIDER_DICT, insert_thres, speed, run_time)
    info.append(len(subsidy_offer))
    info += subsidy_offer_count
    #master_info[data_index].append(info)
    # Basic_class.DataSaver4_summary([info])
    f = open('res/' + data[2] + str(ite) + "mean_" + str(mean_list[ite]) + "std_" + str(
        std_list[ite]) + "_ite_info_save.txt", 'a')
    for ele in info:
        f.write(str(ele) + '\n')
    f.close
    CustomerDataSaver(CUSTOMER_DICT,
                                   'res/' + add_info + data[2] + str(ite) + "mean_" + str(mean_list[ite]) + "std_" + str(
                                       std_list[ite]) + "_ct_save")
    return info
