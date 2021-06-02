# Simulate a Matern point process on a rectangle.
# Author: H. Paul Keeler, 2018.
# Website: hpaulkeeler.com
# Repository: github.com/hpaulkeeler/posts
# For more details, see the post:
# hpaulkeeler.com/simulating-a-matern-cluster-point-process/

import numpy as np  # NumPy package for arrays, random number generation, etc
import matplotlib.pyplot as plt  # For plotting
from openpyxl import Workbook

plt.close('all')  # close all figures

# Simulation window parameters
xMin = -.5
xMax = .5
yMin = -.5
yMax = .5

# Parameters for the parent and daughter point processes
lambdaParent = 4 # density of parent Poisson point process
lambdaDaughter = 120  # mean number of points in each cluster
radiusCluster = 0.1  # radius of cluster disk (for daughter points)

# Extended simulation windows parameters
rExt = radiusCluster  # extension parameter -- use cluster radius
xMinExt = xMin - rExt
xMaxExt = xMax + rExt
yMinExt = yMin - rExt
yMaxExt = yMax + rExt
# rectangle dimensions
xDeltaExt = xMaxExt - xMinExt
yDeltaExt = yMaxExt - yMinExt
areaTotalExt = xDeltaExt * yDeltaExt  # area of extended rectangle

# Simulate Poisson point process for the parents
numbPointsParent = np.random.poisson(areaTotalExt * lambdaParent)  # Poisson number of points

# x and y coordinates of Poisson points for the parent
xxParent = xMinExt + xDeltaExt * np.random.uniform(0, 1, numbPointsParent)
yyParent = yMinExt + yDeltaExt * np.random.uniform(0, 1, numbPointsParent)

# Simulate Poisson point process for the daughters (ie final poiint process)
numbPointsDaughter = np.random.poisson(lambdaDaughter, numbPointsParent)
numbPoints = sum(numbPointsDaughter)  # total number of points
# Generate the (relative) locations in polar coordinates by
# simulating independent variables.
theta = 2 * np.pi * np.random.uniform(0, 1, numbPoints)  # angular coordinates
rho = radiusCluster * np.sqrt(np.random.uniform(0, 1, numbPoints))  # radial coordinates

# Convert from polar to Cartesian coordinates
xx0 = rho * np.cos(theta)
yy0 = rho * np.sin(theta)

# replicate parent points (ie centres of disks/clusters)
xx = np.repeat(xxParent, numbPointsDaughter)
yy = np.repeat(yyParent, numbPointsDaughter)

# translate points (ie parents points are the centres of cluster disks)
xx = xx + xx0
yy = yy + yy0

# thin points if outside the simulation window
booleInside = ((xx >= xMin) & (xx <= xMax) & (yy >= yMin) & (yy <= yMax))
# retain points inside simulation window
xx = xx[booleInside]
yy = yy[booleInside]

rev_x = []
rev_y = []
for x in xx:
    rev_x.append((x + 0.5)*50)
for y in yy:
    rev_y.append((y+0.5)*50)
print('# Parents :', numbPointsParent, '/# Daugters', numbPointsDaughter)
print('After revise #', len(rev_x))

#Save data to csv
file_name = 'MaternCluster loc_data.xlsx'
wb = Workbook()
wb['Sheet'].title = 'loc'
sheet_sm = wb['loc']
row = 1
header = ['Num', 'x', 'y']
for col in range(1, len(header) + 1):
    sheet_sm.cell(row=row, column=col).value = header[col - 1]
row += 1
infos = []
for index in range(len(rev_x)):
    infos.append([index, rev_x[index], rev_y[index]])

for info in infos:
    col = 1
    for ele in info:
        sheet_sm.cell(row=row, column=col).value = ele
        col += 1
    row += 1
wb.save(filename=file_name)


# Plotting
#plt.scatter(xx, yy, edgecolor='b', facecolor='none', alpha=0.5)
plt.scatter(rev_x, rev_y, edgecolor='b', facecolor='none', alpha=0.5)
plt.xlabel('x')
plt.ylabel('y')
plt.axis('equal')
plt.show()
