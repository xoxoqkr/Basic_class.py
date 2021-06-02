# -*- coding: utf-8 -*-
import random
import time
import simpy
import math
import operator
import numpy as np
from openpyxl import Workbook
from datetime import datetime
from sklearn.cluster import KMeans
import copy

class Customer(object):
    def __init__(self, env, name, input_location =  None, end_time = 120 ,service_time = 3):
        self.name = name  # 각 고객에게 unique한 이름을 부여할 수 있어야 함. dict의 key와 같이
        self.time_info = [round(env.now,2), None, None, None, None, end_time , service_time]
        # [0 :발생시간, 1: 차량에 할당 시간, 2:차량에 실린 시간, 3:목적지 도착 시간, 4:고객이 받은 시간, 5: 보장 배송 시간]
        self.locaton = input_location
        self.assigned = False
        self.loaded = False
        self.done = False
        self.cancelled = False
        self.server_info = []
        env.process(self.Decline(env))

    def Decline(self, env):
        yield env.timeout(self.time_info[5])
        if self.assigned == False and self.done == False:
            self.cancelled = True

class CoupangFlex(object):
    def __init__(self, env, name, input_location , expected_fee, customer_set, add_range = 20,  speed = 2, thres = 30,  accept_pr = 0.9):
        self.CP = simpy.Resource(env, capacity=1)
        self.name = name
        self.location = input_location
        self.Done = [False, False]
        self.withdraw = False
        self.assigned = None
        self.fee = expected_fee
        self.gen_time = env.now
        self.thres = thres
        self.add_range = add_range #차고지로 부터 고객까지의 이동시간, add_range이내면 배송을 수행함.
        self.served = []
        self.accept_pr = accept_pr
        self.speed = speed
        self.left = False
        env.process(self.Left(env))


    def CF_Running(self, env, customer, end = True):
        self.Done[0] = True
        self.assigned = env.now
        if customer.time_info[1] != None:
            print ("Error2", customer.name, customer.time_info)
        customer.time_info[1] = env.now
        with self.CP.request() as req:
            customer.assigned = True
            yield req
            moving_duration1 = (distance(self.location, depot.location) / self.speed)
            yield env.timeout(moving_duration1)
            customer.time_info[2] = env.now
            customer.loaded = True
            moving_duration2 = (distance(depot.location, customer.location) / self.speed)
            yield env.timeout(moving_duration2)
            customer.time_info[3] = env.now
            service_time = customer.service_time  # CP의 숙련도에 따라 서비스 시간 차이 생김
            yield env.timeout(service_time)
            customer.time_info[4] = env.now
            customer.done = True
            customer.server_info.append(["C", self.name, self.fee])
            if end == True:
                self.Done[1] = True

    def Left(self, env):
        yield env.timeout(self.thres)
        if self.done[0] == False:
            self.left = True

class Driver(object):
    def __init__(self, env, name, speed):
        self.name = name
        self.route = []  # [[유형,이름,위치,도착시간,출발시간,종료시점,상태유형],...,] 유형 : 0(창고)/1(고객) ; 상태유형 : 0 서비스 X /1 : 서비스 됨
        self.call_back_time = [[0, 'begin']]
        self.veh = simpy.Resource(env, capacity=1)
        self.last_location = [25,25] #송파 집중국
        self.served = [0]
        self.call_back_info = [["start",True,[0,0],[]]]
        self.speed = speed

    def Driving(self, env, customer, customer_set ,cp_num = 2 ,rev_beta = 0 , next_coord=None, fake_parameter=False):
        customer.loaded = True
        customer.time_info[1] = env.now
        customer.assigned = True
        with self.veh.request() as req:
            req.info = customer.name
            yield req  # users에 들어간 이후에 작동
            customer.time_info[2] = env.now
            moving_duration = distance(self.last_location, customer.location) / self.speed
            yield env.timeout(moving_duration)
            customer.time_info[3] = env.now
            yield env.timeout(customer.time_info[6]) #service_time
            customer.time_info[4] = env.now
            self.served.append(customer.name)
            self.last_location = customer.location
            customer.done = True
            if customer.name == 0 and env.now > 5:
                if len(self.veh.put_queue) > 0: #아직 서비스 받지 못한 고객들
                    self.EnrouteRouting()
                    self.call_back_info[-1] #수정
                else:
                    self.IdleThanRun(env, customer_set, cp_num, rev_beta=rev_beta)
                    self.call_back_info.append(["idle", True, [int(env.now), int(env.now)], None])


    def EnrouteRouting(self, env, customer_set, infos, now_time , slack = 15):
        loaded_cts_name = []
        for request in self.veh.put_queue:
            loaded_cts_name.append(request.info)
        ava_cts = []
        for ct in UnloadedCustomer(customer_set, now_time):
            depot2ct = distance(customer_set[0].location, ct.location)/self.speed
            if ct.time_info[5] + ct.time_info[0] >= now_time + depot2ct + slack:
                ava_cts.append(ct.name)
        #NN으로 라우트 형성
        nn_route = NN_Algo(ava_cts + loaded_cts_name, customer_set)
        #2-opt
        nn_route2 = []
        for ct in nn_route: #현재 실려 있는 고객들 U 새로운 고객들로 구성된 적당한 라우트를 만들것.
            nn_route2.append(customer_set[ct])
        nn_route3 = two_opt_solver(nn_route2)
        for ct in nn_route3:
            env.process(self.Driving(self, env, customer_set[ct], customer_set))


    def ToDepot(self, env, customer_set, timing = 0):
        later_ct_names = []
        for request in self.veh.put_queue[timing:]:
            later_ct_names.append(request.info)
        self.veh.put_queue = self.veh.put_queue[:timing]
        env.process(self.Driving(env, customer_set[0], customer_set))
        self.call_back_info.append(['callback', False, [env.now, None],later_ct_names])
        print("ToDepot::", timing ,later_ct_names)

    def IdleThanRun(self, env, customer_set, cp_num, rev_beta = None, given_cts = None):
        unas_cts = []
        unas_cts_names = []
        for ct_key in customer_set:
            ct = customer_set[ct_key]
            if ct.assigned == False and ct.time_info[1] == None:
                unas_cts.append(ct)
                unas_cts_names.append(ct.name)
        route = two_opt_solver(unas_cts)
        removed_ct_names = Allocate2CP(route, customer_set, cp_num, benefit_beta=0)
        assigned_cts = [item for item in unas_cts_names if item not in removed_ct_names]
        validated_cts, assigned_cts = self.ValidationCheck(self, assigned_cts, customer_set, env.now)
        for ct in assigned_cts:
            env.process(self.Driving(self, env, customer_set[ct], customer_set))
        print("IdleThanRun::validated_cts #", len(validated_cts), "assigned_cts #", len(assigned_cts))


    def ValidationCheck(self, customers, customer_set, now_time):
        assigned_cts = []
        validated_cts = []
        time = now_time
        modified_index = 0
        if customers[0] != 0:
            time += distance(customer_set[0].location, customer_set[customers[0]].location)/self.speed
        for index in range(1,len(customers)):
            bf_ct = customer_set[customers[min(modified_index,index - 1)]]
            this_ct = customer_set[customers[index]]
            added_time = distance(bf_ct.location, this_ct.location)/self.speed
            if time + added_time <= this_ct.time_info[0] + this_ct.time_info[5] :
                assigned_cts.append(this_ct.name)
                time += added_time + this_ct.time_info[6]
            else:
                validated_cts.append(this_ct.name)
                modified_index = index - 1
        return validated_cts, assigned_cts


def CustomerGenerator(env, customer_list ,dir = None):
    datas = open(dir, 'r')
    lines = datas.readlines()
    for line in lines[1:]:
        data = line.split(';')
        if float(data[4]) == 0:
            c = Customer(env, int(data[0]), input_location = [float(data[1]), float(data[2])], end_time = 1000)
        else:
            c = Customer(env, int(data[0]), input_location = [float(data[1]), float(data[2])])
        customer_list.append(c)
        yield env.timeout(float(data[3]))

def RouteTime(route):
    res = 0
    for index in range(1, len(route)):
        res += distance(route[index - 1], route[index])
    return round(res, 4)

def SingleNodeInsertCost(org_left,org_right,inserted):
    org = distance(org_left, org_right)
    rev = distance(org_left, inserted) + distance(inserted, org_right)
    return rev , org


def CFMaker(env, CoupangFlex_set ,expected_fee_l, expected_fee_u, Fee_list, customer_set,fx = 'lin2'):
    name = 0
    while True:
        expected_fee = random.randrange(expected_fee_l ,expected_fee_u ,step = 100)
        if fx == 'lin1':# Todo: CP를 지불 금액에 따른 발생률로 바꿀 것.
            lamda = 20/(3000/expected_fee)
        elif fx == 'lin2':
            lamda = (9.0 / 200) * expected_fee - 80
        elif fx == 'exp1' :
            lamda = 60.0/(15**(1000/expected_fee))
        else: #fx = 'exp2'
            lamda = 2**(600/expected_fee)
        Fee_list.append([expected_fee, env.now])
        ob = CoupangFlex(env, name, expected_fee, customer_set)
        CoupangFlex_set.append(ob)
        name += 1
        lamda = random.randrange(5, 10)
        yield env.timeout(lamda)


def distance(x1, x2, real_cor = False):
    if real_cor == False:
        return round(math.sqrt((x1[0] - x2[0]) ** 2 + (x1[1] - x2[1]) ** 2), 2)
    else:
        return round(coord_cal.distance_r(x1,x2),2)

#  TODO : 고객을 생성된 라우트에서 제거하는 과정
def Allocate2CP(route_seq, customer_infos ,cp_num, benefit_beta = 0, info_parammeter = False):
    #route_seq : 고객 이름으로 이루어진 리스트
    for _ in range(0, cp_num):  # 현재 가능한 CP만큼 고객을 제거 해 보자.
        if random.random() < 0.5:
            del_infos = Few_longest_node_set(route_seq, customer_infos ,cp_num, benefit_beta = benefit_beta)
        else:
            del_infos = Worst_node_set(route_seq, customer_infos, cp_num, benefit_beta = benefit_beta)
    del_infos.sort(key=operator.itemgetter(2))
    remove_ct_names = []
    remove_ct_names2 = []
    for index in range(0,min(cp_num, len(del_infos))):
        if del_infos[index][1] >= benefit_beta:
            remove_ct_names.append(del_infos[index][0])
        if info_parammeter == True:
            remove_ct_names2.append([del_infos[index][0], del_infos[index][2]]) #[[이름, 비용],...]
    if info_parammeter == False:
        return remove_ct_names
    else:
        return remove_ct_names2

def Few_longest_node_set(route_seq, customer_infos ,cp_num, benefit_beta=0, route_para=True):
    del_infos = []
    for ct_name in route_seq:
        del_infos.append([ct_name, distance(depot.location, customer_infos[ct_name].location), customer_infos[ct_name].time_info[0]])
    del_infos.sort(key=operator.itemgetter(1), reverse=True)
    return del_infos


def Worst_node_set(route_seq, customer_infos ,cp_num, benefit_beta=0):
    del_infos = []
    route_cost = 0
    for node_index in range(1,len(route_seq)):
        route_cost += distance(customer_infos[route_seq[node_index - 1]].location,customer_infos[route_seq[node_index]].location)
    for target_node in route_seq[1:-1]:
        if target_node != 0:
            tar = customer_infos[route_seq[route_seq.index(target_node)]].location
            bf = customer_infos[route_seq[route_seq.index(target_node) - 1]].location
            af = customer_infos[route_seq[route_seq.index(target_node) + 1]].location
            rev_cost = route_cost - distance(bf,tar) - distance(tar, af) + distance(bf, af)
            del_infos.append([target_node, rev_cost, customer_infos[route_seq[route_seq.index(target_node)]].time_info[0]])
    return del_infos

#TODO : 차량 라우트 구성 방법

def two_opt2(route, cost_mat, env, max_cal_time = 60):
    best = route
    improved = True
    t1 = datetime.now()
    while improved  and (datetime.now()-t1).seconds <= max_cal_time:
        improved = False
        for i in list(range(1, len(route) - 2)):
            for j in list(range(i + 1, len(route))):
                if j - i == 1: continue
                if cost_change(cost_mat, best[i - 1], best[i], best[j - 1], best[j]) < 0:
                    if best[i] == 0:
                        pass
                    else:
                        best[i:j] = best[j - 1:i - 1:-1]
                        improved = True
        route = best
    print("Now:",env.now,"/For # ct:",len(route),"/two opt2 time:" ,(datetime.now()-t1).seconds, "sec" )
    return best

def cost_change(cost_mat, n1, n2, n3, n4):
    return cost_mat[n1][n3] + cost_mat[n2][n4] - cost_mat[n1][n2] - cost_mat[n3][n4]


def two_opt_solver(route, max_cal_time = 60):
    route_seq = route
    dist_mat = np.zeros((len(route_seq),len(route_seq)))
    row_count = 0
    for i in route_seq:
        col_count = 0
        for j in route_seq:
            if col_count > row_count :
                dist_mat[row_count, col_count] = distance(i.location, j.location)
            col_count += 1
        row_count += 1
    dist_mat += dist_mat.T
    dist_mat = list(dist_mat)
    route_name_seq = range(0,len(route))
    best_route = two_opt2(route_name_seq, dist_mat, env)
    org_name_route = []
    for i in best_route:
        org_name_route.append(route[i].name) #org_name_route = [고객이름, 고객이름 ,... , 고객이름]
    return org_name_route

def InputNamesReturnClass(names, class_set):
    res = []
    for name in names:
        res.append(class_set[name])
    return res

def UnloadedCustomer(customer_set, now_time):
    res = []
    for ct_name in customer_set:
        customer = customer_set[ct_name]
        cond1 = now_time - customer.time_info[0] < customer.time_info[5]
        cond2 = customer.assigned == False
        cond3 = customer.done == False
        cond4 = customer.time_info[1] == None
        if cond1 and cond2 and cond3 and cond4 == True:
            res.append(customer)
    return res

def RouteWithLoadedCustomers(loaded_customer_names, customer_set):
    loaded_customers = InputNamesReturnClass(loaded_customer_names, customer_set)
    load_route = two_opt_solver(loaded_customers)

def NN_Algo(ct_names, customer_set):
    res = [0]
    del_ct = []
    for ct1 in ct_names:
        from_ct = customer_set[res[-1]]
        tem_dist = []
        for ct2 in ct_names:
            if ct2 != ct1 and ct2 not in res:
                tem_dist.append([ct2, distance(customer_set[ct1].location,customer_set[ct2].location)])
        tem_dist.sort(key = operator.itemgetter(1))
        res.append(tem_dist[0][0])
    return res


def NNearestRouteMaker(env, driver_set, customer_set):
    data = {}
    if len(driver_set) > 2:
        data = clustered_coord(driver_set, customer_set)
    else:
        data[0] = customer_set
    index = 0
    for key in data:
        two_opt_route = []
        for node in data[key]:
            ct = node
            two_opt_route.append(ct)
        route = two_opt_solver(two_opt_route)
        for ct_name in route:
            if ct_name != 0:
                env.process(driver_set[index].Driving(env, customer_set[ct_name], customer_set))
        if route[-1] != 0:
            env.process(driver_set[index].Driving(env, depot, customer_set))
        index += 1

def clustered_coord(driver_set, customer_set):
    if len(customer_set) < 2:
        print ("clustered_coord : less than 2 elements")
        return None
    node_coor = []
    for ct in customer_set:
        node_coor.append(customer_set[ct].location)
    node_coor = np.asarray(node_coor)
    kmeans = KMeans(n_clusters=len(driver_set), random_state=0).fit(node_coor)
    cluster_assigned = {}
    for index in list(set(kmeans.labels_)):
        cluster_assigned[index] = []
    #print("clustered info", list(set(kmeans.labels_)), cluster_assigned)
    #print(kmeans.labels_)
    #print(customer_set)
    index = 0
    for clt in kmeans.labels_:
        cluster_assigned[clt].append(customer_set[index])
        index += 1
    return cluster_assigned

#TODO : 회차 판별기

def CustomerResetter(customer, now_time):
    for index in range(1,5):
        customer.time_info[index] = None
    customer.reset_info.append([now_time, "reset"])


def ReturnChecker(env, driver_set, customer_set, ramda_record_list ,cumul_num = 15, CPs = None):
    unserved_ct = UnloadedCustomer(customer_set, env.now)
    print(env.now, "UnAs CTs#", len(unserved_ct))
    if len(unserved_ct) > 20:
        return_score = []
        for driver_name in driver_set:
            driver = driver_set[driver_name]
            condition1 = driver.call_back_info[-1][2][1] <= env.now - 90
            condition2 = ExpectedCustomerLamda(ramda_record_list, env.now, env.now + 90) > cumul_num #TODO : 조건 수정하기
            # return Timing 확인
            tem = ReturnTiming(driver, env.now)
            if condition2 == True:
                print("return condition checked// T From",env.now,"~", env.now + 90)
                pass
            if condition1 or condition2 == True:
                remain_cts = len(driver.veh.users) + len(driver.veh.put_queue)
                return_score.append([driver.name, remain_cts])
        return_score.sort(key = operator.itemgetter(1))
        if len(return_score) > 0:#차량 회차 실시
            return_driver = driver_set[return_score[0][0]] # 지금은 바로 당장 시행
            return_cost = []
            for request in return_driver.veh.put_queue:
                if request.info.name > 0:
                    dist2depot = distance(request.info.location, customer_set[0].location)
                    return_cost.append(dist2depot)
            if len(return_cost) > 0:
                cond1 = return_driver.call_back_info[-1][0] == "callback"
                cond2 = return_driver.call_back_info[-1][1] == False
                cond3 = return_driver.call_back_info[-1][0] == "start"
                if cond1 and cond2 == False or cond3 == True:
                    print(env.now, "driver#", return_driver.name, " return")
                    return_point = return_cost.index(min(return_cost))
                    return_driver.ToDepot(env, customer_set, timing = return_point)
        else:
            print(env.now, "No return S")
    else:
        print(env.now, "No return L")
    yield env.timeout(5)

def ExpectedCustomerLamda(data, now, end):
    now_h = now//60
    end_h = end//60
    res = []
    for index in data:
        if now_h <= index < end_h:
            res.append(data[index][0])
    #print("ExpectedCustomerLamda", now, end,res,now_h, end_h)
    return sum(res)

def ReturnTiming(veh, now):
    print("veh",veh.name, "check")
    if len(veh.veh.users) > 0:
        seq = veh.veh.users + veh.veh.put_queue
        res = []
        mid_time = 0
        for index in range(1,len(seq)) :
            bf = seq[index - 1].info
            af = seq[index].info
            rev, org = SingleNodeInsertCost(bf.location, af.location, depot.location)
            Todepot = distance(bf.location,depot.location)
            customer_end_time = (af.time_info[0] + af.time_info[5] + mid_time)
            val1 = round((customer_end_time - (rev + now))/speed,1)
            #val1 = 고객 j의 종료시점 - 수정된 이동 경로 종료 시점 -> 차고지를 왕복한 이후의 여유 시간.
            val2 = round((customer_end_time - (Todepot + now))/speed,1)
            #val2 = 고객 j의 종료시점 - 수정된 이동 경로 종료 시점-> 차고지 도착 후의 여유 시간.
            #proposition : val1 > val2
            if val2 < 0:
                res.append([index, val1, val2])
            mid_time += org
            tem_res = []
            mid_time2 = 0
            for index2 in range(index + 1,len(seq)) :
                bf1 = seq[index2 - 1].info
                af1 = seq[index2].info
                customer_end_time2 = (af1.time_info[0] + af1.time_info[5])
                move_time = distance(bf1.location,af1.location)/speed

                val1_tem = round(customer_end_time2 - (rev + move_time + now + mid_time2),1)
                val2_tem = round(customer_end_time2 - (Todepot + move_time + now + mid_time2),1)
                if val2_tem < 0:
                    tem_res.append([index2, val1_tem, val2_tem])
                mid_time2 += move_time
            if len(tem_res) > 0:
                print(index,"// violated num:", len(tem_res),"/",len(seq) - index + 1, "//", tem_res)
        res.sort(key = operator.itemgetter(2), reverse = True)
        print("remaining Time//ava num:",len(seq) + 1 - len(res) ,"//",res[:3])
        return res
    else:
        return None

def Return_Score(veh_set, customer_set, now_time):
    unloaded_customer = UnloadedCustomer(customer_set, now_time)
    print(now_time, "current unloaded ct#", len(unloaded_customer))
    for veh_index in veh_set:
        veh = veh_set[veh_index]
        orders = veh.veh.users + veh.veh.put_queue
        route_length = []
        for index in range(1,len(orders)):
            route_length.append(distance(orders[index - 1].info.location, orders[index].info.location))
        ave_len = 0
        if len(route_length) > 0:
            ave_len = sum(route_length)/len(route_length)
            print("veh:",veh_index,"/Len: ",len(orders), "/ dist2depot:",distance(orders[0].info.location, customer_set[0].location),"/Ave dist:",ave_len)
        else:
            print("veh:", veh_index, "/Len: ", 0, "/ dist2depot:", 0, "/Ave dist:", ave_len)
#TODO : 플렉스 할당 판별기

def PlatformRunner(env, driver_set, customer_set,ramda_record_list, cf_set = None):
    #차량의 초기해 구성 후 출발
    NNearestRouteMaker(env, driver_set, customer_set) #NN 알고리즘으로 초기해 구성
    yield env.timeout(30)
    while True:    #운행이 계속 되는 동안 차량의 회차 및 플렉스 할당 결정
        """
        veh_num = random.randrange(len(driver_set))
        print("call back", veh_num)
        driver_set[veh_num].ToDepot(env, customer_set, timing = 3)
        index += 1        
        """
        env.process(ReturnChecker(env, driver_set, customer_set, ramda_record_list))
        Return_Score(driver_set, customer_set, env.now)
        yield env.timeout(5)

def Sub_problem_solver(env, veh_set, customer_set, nowtime):
    yield env.timeout(30)
    gl_res = []
    for veh_name in veh_set:
        veh = veh_set[veh_name]
        res = "recall possible index list"
        gl_res.append(res)

    pass



def Summary(driver_set, customer_set, CoupangFlex_set):
    res = []
    res_v = []
    res_cf = []
    for ct_ke in customer_set:
        customer= customer_set[ct_ke]
        if customer.done == True:
            if customer.time_info[4] == None:
                print(customer.name,"TimeInfo",customer.time_info)
            else:
                res.append(customer.time_info[4] - customer.time_info[0])
                if customer.mode[0] == 'V':
                    res_v.append(customer.time_info[4] - customer.time_info[0])
                else:
                    res_cf.append(customer.time_info[4] - customer.time_info[0])
    CF_res = []
    for CF in CoupangFlex_set:
        if len(CF.served) > 0:
            CF_res.append(len(CF.served))
    print ("Total", len(customer_set))
    print ("Total served", len(res), "Ave : ", sum(res)/len(res))
    print ("Veh served", len(res_v), "Ave : ", sum(res_v) / len(res_v))
    if len(res_cf) > 0 and len(CF_res) > 0:
        print ("CF served", len(res_cf), "Ave : ", sum(res_cf) / len(res_cf), "paid fee", len(res_cf)*900, "Served Ave", sum(CF_res)/len(CF_res))

#data recorder
def DataSaveAsXlxs(veh_set, CPs, customer_set):
    now = datetime.today().isoformat()
    date_info_index = [[2, 4], [5, 7], [8, 10], [14, 16], [17, 19], [21, 23]] # YYMMDDTTMMSS
    date_time = str()
    for index in date_info_index:
        date_time += now[int(index[0]):int(index[1])]
    file_name = date_time + 'res.xlsx'
    wb = Workbook()
    wb['Sheet'].title = 'veh'
    sheet_veh = wb['veh']
    header_veh = ['No.', 'Served_ct_#', 'CB t','Remain_queue','call_back_info','route_cost','served_seq']
    for col in range(1, len(header_veh) + 1):
        sheet_veh.cell(row = 1, column = col).value = header_veh[col - 1]
    for veh_index in veh_set:
        r_cost = 0
        veh = veh_set[veh_index]
        for ct_name in veh.served[1:]:
            bf = customer_set[ct_name].location
            af = customer_set[ct_name].location
            r_cost += distance(bf,af)/speed
        info = [veh.name , len(veh.served), str(veh.call_back_time), len(veh.veh.users)  + len(veh.veh.put_queue),r_cost,str(veh.call_back_info), str(veh.served)]
        ele_count = 1
        start_row = sheet_veh.max_row + 1
        for ele in info:
            sheet_veh.cell(row = start_row, column = ele_count).value = ele
            ele_count += 1
    header_dr = ['No.', 'used #', 'declined #', 'Ave waiting time']
    wr_row = sheet_veh.max_row +2
    for col in range(1, len(header_dr) + 1):
        sheet_veh.cell(row = wr_row, column = col).value = header_dr[col - 1]
    cp1 = 0
    cp2 = 0
    cp3 = 0
    cp4 = []
    if cp2 > 0:
        for cp in CPs:
            cp1 += 1
            if cp.Done == True:
                cp2 += 1
                cp4.append(cp.assigned)
            else:
                cp3 += 1
        dr_data = [cp1,cp2,cp3,sum(cp4)/cp2]
        wr_row += 1
        for col in range(1, len(dr_data) + 1):
            sheet_veh.cell(row= wr_row, column=col).value = dr_data[col - 1]

    header_dr = ['Served #.', 'Ave wating time']
    wr_row = sheet_veh.max_row + 2
    for col in range(1, len(header_dr) + 1):
        sheet_veh.cell(row = wr_row, column = col).value = header_dr[col - 1]
    ltd_list = []
    for ct_key in customer_set:
        ct = customer_set[ct_key]
        if ct.time_info[4] != None:
            ltd = ct.time_info[4] - ct.time_info[0]
            ltd_list.append(ltd)
    dr_data = [len(ltd_list), sum(ltd_list) / len(ltd_list)]
    wr_row += 1
    for col in range(1, len(dr_data) + 1):
        sheet_veh.cell(row=wr_row, column=col).value = dr_data[col - 1]

    sheet_ct = wb.create_sheet('ct')
    #sheet_ct = wb2.active
    #sheet_ct.title = 'ct'
    header_ct = ['No.', 'Occurred', 'Assigned', 'Loaded', 'Arrived', 'Served','CP_wait','CP_fee','Mode','cor_x','cor_y','reset_info']
    # [발생시간, 차량에 할당 시간, 차량에 실린 시간, 목적지 도착 시간, 고객이 받은 시간]
    for col in range(1, len(header_ct) + 1):
        sheet_ct.cell(row = 1, column = col).value = header_ct[col - 1]
    for key in customer_set:
        info = [key]
        for time in customer_set[key].time_info[:5]:
            info.append(time)
        if customer_set[key].CP_info[1] != None:
            info.append(customer_set[key].time_info[1] - customer_set[key].CP_info[1])
            info.append(customer_set[key].CP_info[2])
        else:
            info.append('None')
            info.append('None')
        #info.append(customer_infos[key].CP_info[2])
        info.append(customer_set[key].mode)
        info.append(customer_set[key].location[0])
        info.append(customer_set[key].location[1])
        info.append(str(customer_set[key].reset_info))
        ele_count = 1
        start_row = sheet_ct.max_row + 1
        for ele in info:
            sheet_ct.cell(row = start_row, column = ele_count).value = ele
            ele_count += 1
    """
    sheet_beta = wb.create_sheet('rev_beta') #E_계열의 함수를 확인하기 위한 sheet
    start_row = sheet_beta.max_row + 1
    ele_count = 1
    header_sheet_beta = ['time', 'rev_beta', 'ct_num', 'expected_fee', 'lamda gen ct #']
    # [발생시간, 차량에 할당 시간, 차량에 실린 시간, 목적지 도착 시간, 고객이 받은 시간]
    for col in range(1, len(header_sheet_beta) + 1):
        sheet_beta.cell(row = 1, column = col).value = header_sheet_beta[col - 1]
    for beta in rev_beta_infos:
        start_t = (start_row - 1) * 30
        end_t = start_row * 30
        sheet_beta.cell(row=start_row, column=ele_count).value = start_t
        sheet_beta.cell(row=start_row, column=ele_count + 1).value = beta
        ct_count = 0
        for key in customer_set:
            ct = customer_set[key]
            if start_t <= ct.time_info[0]  < end_t:
                ct_count += 1
            elif ct.time_info[0]  > end_t:
                break
        sheet_beta.cell(row=start_row, column=ele_count + 2).value = ct_count
        target_fee = []
        for fee in Fee_list:
            if start_t <= fee[1] < end_t:
                target_fee.append(fee[0])
            elif fee[1] > end_t:
                break
        sheet_beta.cell(row=start_row, column=ele_count + 3).value = round(sum(target_fee)/len(target_fee),2)
        for lamda_info in lamda_infos_list:
            sheet_beta.cell(row=start_row, column=ele_count + 4).value = lamda_info[0]
        start_row += 1
    """
    sheet_cf = wb.create_sheet('CF')  # E_계열의 함수를 확인하기 위한 sheet
    header_sheet_cf = ['num#', 'occur_time', 'assigned_time', 'served_ct_info']
    for col in range(1, len(header_sheet_cf) + 1):
        sheet_cf.cell(row=1, column=col).value = header_sheet_cf[col - 1]
    # [발생시간, 차량에 할당 시간, 차량에 실린 시간, 목적지 도착 시간, 고객이 받은 시간]
    row = 2
    for cf in CPs:
        sheet_cf.cell(row = row, column=1).value = str(cf.name)
        sheet_cf.cell(row = row, column=2).value = str(cf.time)
        sheet_cf.cell(row = row, column=3).value = str(cf.assigned)
        sheet_cf.cell(row = row, column=4).value = str(cf.ct_info)
        row += 1
    wb.save(filename=file_name)



